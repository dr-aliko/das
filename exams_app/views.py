import json
from datetime import date, timedelta
from types import SimpleNamespace
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.db import models
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_http_methods

from users_app.decorators import coach_can_view_student, coach_required, student_required

from .forms import BransDenemeForm
from .models import BransDeneme, BransTopicError, Exam, ExamResult, ExamTopicError, Publisher, Subject, StudentTask, Topic

SUBJECT_COLORS = {
    'TYT Türkçe':         {'border': '#3b82f6', 'background': 'rgba(59,130,246,0.15)'},
    'TYT Matematik':      {'border': '#8b5cf6', 'background': 'rgba(139,92,246,0.15)'},
    'TYT Sosyal Bilimler':{'border': '#f59e0b', 'background': 'rgba(245,158,11,0.15)'},
    'TYT Fen Bilimleri':  {'border': '#10b981', 'background': 'rgba(16,185,129,0.15)'},
}
DEFAULT_COLORS = [
    {'border': '#6366f1', 'background': 'rgba(99,102,241,0.15)'},
    {'border': '#ec4899', 'background': 'rgba(236,72,153,0.15)'},
]

# All subject names that contribute to the "Fen Bilimleri" analytics line
# (old umbrella entry + new split subjects added for BransDeneme)
_FEN_NAMES = frozenset({'TYT Fen Bilimleri', 'TYT Fizik', 'TYT Kimya', 'TYT Biyoloji'})
# Canonical four-subject axis shown in every chart / comparison / analytics view
_TYT_DISPLAY_SUBJECTS = ['TYT Türkçe', 'TYT Matematik', 'TYT Sosyal Bilimler', 'TYT Fen Bilimleri']


# ──────────────────────────────────────────────
# STUDENT — DASHBOARD v2 helpers (DAS-008/009/010)
# ──────────────────────────────────────────────

_V2_SUBJECT_MAP = [
    ('TYT Türkçe',          'turkce', 'Türkçe'),
    ('TYT Matematik',       'mat',    'Matematik'),
    ('TYT Sosyal Bilimler', 'sosyal', 'Sosyal'),
    ('TYT Fen Bilimleri',   'fen',    'Fen'),
]

def _v2_period_label(period):
    return {'30': '30 gün', '90': '3 ay'}.get(period, 'Tümü')


_TR_MONTHS_SHORT = ['Oca','Şub','Mar','Nis','May','Haz','Tem','Ağu','Eyl','Eki','Kas','Ara']

def _fmt_task_date(dt):
    """Format a datetime as '12 May' in Turkish for task lifecycle display."""
    if not dt:
        return None
    return f'{dt.day} {_TR_MONTHS_SHORT[dt.month - 1]}'

def _task_cycle_days(created_at, completed_at):
    """Days from task creation to completion (0 = same day)."""
    if not (created_at and completed_at):
        return None
    return max(0, (completed_at.date() - created_at.date()).days)

def _build_v2_stats(exam_list, period):
    count = len(exam_list)
    if not count:
        return {
            'avg_net': 0, 'avg_net_delta': None,
            'last_net': 0, 'last_net_delta': None,
            'max_net': 0, 'net_trend': None,
            'exam_count': 0, 'last_exam_title': '',
            'last_exam_date': '', 'selected_period': _v2_period_label(period),
        }
    # Exam.total_net is currently a method, but tolerate @property / field
    # without forcing a model change. Django templates already auto-call methods.
    nets = [float(e.total_net() if callable(e.total_net) else e.total_net) for e in exam_list]
    avg_net = round(sum(nets) / count, 2)
    last_net = nets[0]
    last_net_delta = round(last_net - nets[1], 2) if count > 1 else None
    recent = nets[:5]
    prior = nets[5:10]
    avg_net_delta = round(sum(recent) / len(recent) - sum(prior) / len(prior), 2) if prior else None
    max_net = round(max(nets), 2)
    net_trend = round(last_net - avg_net, 2) if count > 1 else None
    last = exam_list[0]
    exam_date = last.exam_date
    last_exam_date = exam_date.strftime('%d.%m') if hasattr(exam_date, 'strftime') else str(exam_date)
    return {
        'avg_net': avg_net,
        'avg_net_delta': avg_net_delta,
        'last_net': last_net,
        'last_net_delta': last_net_delta,
        'max_net': max_net,
        'net_trend': net_trend,
        'exam_count': count,
        'last_exam_title': last.custom_name,
        'last_exam_date': last_exam_date,
        'selected_period': _v2_period_label(period),
    }

def _build_v2_subjects(exam_list):
    """Per-subject stats including accuracy, from last two exams' prefetched results.

    'TYT Fen Bilimleri' is not entered as a single result — students enter
    Fizik/Kimya/Biyoloji individually, so Fen is aggregated from _FEN_NAMES.
    """
    if not exam_list:
        return []
    last_results = {r.subject.name: r for r in exam_list[0].results.all()}
    prev_results = {r.subject.name: r for r in exam_list[1].results.all()} if len(exam_list) > 1 else {}
    out = []
    for subj_name, key, label in _V2_SUBJECT_MAP:
        if subj_name == 'TYT Fen Bilimleri':
            # Aggregate from sub-subjects (Fizik, Kimya, Biyoloji)
            fen_subs = _FEN_NAMES - {'TYT Fen Bilimleri'}
            fen_last = [last_results[s] for s in fen_subs if s in last_results]
            if not fen_last:
                continue
            net = round(sum(float(fr.net_score) for fr in fen_last), 2)
            fen_prev = [prev_results[s] for s in fen_subs if s in prev_results]
            delta = round(net - sum(float(fr.net_score) for fr in fen_prev), 2) if fen_prev else None
            correct = sum(fr.correct_answers for fr in fen_last)
            wrong   = sum(fr.wrong_answers   for fr in fen_last)
            blank   = sum(fr.blank_answers   for fr in fen_last)
            total_q = correct + wrong + blank
            accuracy = round(correct / total_q * 100) if total_q > 0 else 0
            out.append({'key': key, 'name': label, 'net': net, 'delta': delta, 'accuracy': accuracy})
        else:
            r = last_results.get(subj_name)
            if r is not None:
                net = float(r.net_score)
                prev = prev_results.get(subj_name)
                delta = round(net - float(prev.net_score), 2) if prev is not None else None
                total_q = r.correct_answers + r.wrong_answers + r.blank_answers
                accuracy = round(r.correct_answers / total_q * 100) if total_q > 0 else 0
                out.append({'key': key, 'name': label, 'net': net, 'delta': delta, 'accuracy': accuracy})
    return out

# ──────────────────────────────────────────────
# STUDENT — DASHBOARD
# ──────────────────────────────────────────────

_TURKISH_MONTHS = {
    1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan',
    5: 'Mayıs', 6: 'Haziran', 7: 'Temmuz', 8: 'Ağustos',
    9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık',
}


_PERIOD_DAYS = {'7': 7, '30': 30, '90': 90, '365': 365}


def _build_denemeler_v2_context(user, filters=None):
    """DAS-303/305-307: builds month-archived exam context with optional filter/sort."""
    from itertools import groupby as _groupby
    filters = filters or {}

    qs = (
        Exam.objects
        .filter(student=user)
        .select_related('publisher')
        .prefetch_related('results')
        .order_by('-exam_date', '-id')
    )

    # has_note filter (DAS-306)
    if filters.get('has_note'):
        qs = qs.filter(student_note__isnull=False).exclude(student_note='')

    exams_raw = list(qs)

    # Annotate with computed net, delta, and type label.
    exam_objs = []
    for e in exams_raw:
        net = round(sum(r.net_score for r in e.results.all()), 2)
        exam_objs.append(SimpleNamespace(
            id=e.id,
            date=e.exam_date,
            net=net,
            delta=None,
            type='TYT',
            publisher=e.publisher.name,
            custom_name=e.custom_name,
        ))

    # delta always computed on date-ordered list (newest-first) regardless of sort
    for i in range(len(exam_objs) - 1):
        exam_objs[i].delta = round(exam_objs[i].net - exam_objs[i + 1].net, 2)

    # Period filter: trim to last N days (DAS-305)
    period_str = filters.get('period', '')
    if period_str in _PERIOD_DAYS:
        period_cutoff = date.today() - timedelta(days=_PERIOD_DAYS[period_str])
        exam_objs = [e for e in exam_objs if e.date >= period_cutoff]

    # Sort (DAS-307) — applied after period trim, before recent/archive split
    sort_by = filters.get('sort', 'date')
    if sort_by == 'net':
        exam_objs.sort(key=lambda e: e.net, reverse=True)
    elif sort_by == 'publisher':
        exam_objs.sort(key=lambda e: e.publisher.lower())
    # 'date' keeps the existing -exam_date,-id order

    # Partition: recent = last 30 days, older = archive
    cutoff = date.today() - timedelta(days=30)
    recent = [e for e in exam_objs if e.date >= cutoff]
    older  = [e for e in exam_objs if e.date < cutoff]

    # Group archive by (year, month), descending
    archive = []
    for (year, month), group in _groupby(older, lambda e: (e.date.year, e.date.month)):
        month_exams = list(group)
        archive.append(SimpleNamespace(
            month_label=f'{_TURKISH_MONTHS[month]} {year}',
            count=len(month_exams),
            exams=month_exams,
        ))

    sparse_threshold = 3
    return {
        'exams': SimpleNamespace(
            recent=recent,
            archive=archive,
            total_count=len(exam_objs),
        ),
        'sparse_threshold': sparse_threshold,
        'exams_needed': max(0, sparse_threshold - len(recent)),
    }


_BRANS_HUB_SUBJECTS = [
    ('mat', 'Matematik', 'var(--mat)', 'accent-mat', {'TYT Matematik'}),
    ('turkce', 'Türkçe', 'var(--turkce)', 'accent-turkce', {'TYT Türkçe'}),
    ('sosyal', 'Sosyal', 'var(--sosyal)', 'accent-sosyal', {'TYT Sosyal Bilimler'}),
    ('fen', 'Fen', 'var(--fen)', 'accent-fen', {'TYT Fen Bilimleri'}),
    ('fizik', 'Fizik', 'var(--fizik)', 'accent-fizik', {'TYT Fizik'}),
    ('kimya', 'Kimya', 'var(--kimya)', 'accent-kimya', {'TYT Kimya'}),
    ('biyoloji', 'Biyoloji', 'var(--biyoloji)', 'accent-biyoloji', {'TYT Biyoloji'}),
]


def _brans_student_initials(student):
    parts = [p for p in (student.full_name or '').split() if p]
    return ''.join(p[0] for p in parts).upper()[:2] or '?'


def _brans_exam_label(entry):
    return f'{entry.tarih.day} {_TURKISH_MONTHS[entry.tarih.month]}'


def _safe_next_url(request, fallback):
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return fallback


def _append_next(url, next_url):
    return f'{url}?{urlencode({"next": next_url})}' if next_url else url


def _brans_subject_bucket(subject_name):
    for key, _label, _color, _klass, names in _BRANS_HUB_SUBJECTS:
        if subject_name in names:
            return key
    return None


def _subject_detail_source_stats(entries, current_cutoff, previous_cutoff):
    current_entries = [e for e in entries if e.date >= current_cutoff]
    previous_entries = [e for e in entries if previous_cutoff <= e.date < current_cutoff]
    current_nets = [float(e.net) for e in current_entries]
    previous_nets = [float(e.net) for e in previous_entries]
    current_net = round(sum(current_nets) / len(current_nets), 1) if current_nets else 0.0
    prev_net = round(sum(previous_nets) / len(previous_nets), 1) if previous_nets else 0.0
    delta = round(current_net - prev_net, 2) if current_nets and previous_nets else 0.0
    best_obj = max(current_entries, key=lambda e: e.net) if current_entries else None
    worst_obj = min(current_entries, key=lambda e: e.net) if current_entries else None
    avg3_entries = current_entries[:3]
    avg10_entries = current_entries[:10]

    return {
        'currentNet': current_net,
        'delta': delta,
        'trend': 'down' if delta < 0 else 'up',
        'best': {'net': float(best_obj.net), 'date': best_obj.date_label} if best_obj else None,
        'worst': {'net': float(worst_obj.net), 'date': worst_obj.date_label} if worst_obj else None,
        'avg3': round(sum(float(e.net) for e in avg3_entries) / len(avg3_entries), 1) if avg3_entries else 0.0,
        'avg10': round(sum(float(e.net) for e in avg10_entries) / len(avg10_entries), 1) if avg10_entries else 0.0,
    }


def _subject_detail_chart_points(entries, current_cutoff):
    return [
        {
            'date': e.date_label,
            'isoDate': e.date.isoformat(),
            'net': float(e.net),
            'source': e.source,
        }
        for e in sorted((e for e in entries if e.date >= current_cutoff), key=lambda item: (item.date, item.id))
    ]


def _brans_period_cutoff(period):
    days = {'7g': 7, '30g': 30, '3 ay': 90, '1 yıl': 365}.get(period)
    return date.today() - timedelta(days=days) if days else date.min


def _brans_export_period_query(period):
    return period if period in {'7g', '30g', '3 ay', '1 yıl', 'Tümü'} else '30g'


def _brans_insight_url(subject_key, *, coach_view=False, student_id=None):
    if coach_view:
        return reverse('coach:student_brans_subject_detail', kwargs={'student_id': student_id, 'subject_slug': subject_key})
    return reverse('brans_subject_detail', kwargs={'subject_slug': subject_key})


def _build_brans_insights(student, entries, period, *, coach_view=False):
    current_cutoff = _brans_period_cutoff(period)
    total_count = len(entries)
    if total_count < 2:
        text = (
            f"{student.full_name} için en az iki branş denemesi eklendiğinde içgörüler oluşacak."
            if coach_view else
            'En az iki branş denemesi eklediğinde içgörüler burada görünecek.'
        )
        return [{
            'text': text,
            'severity': 'empty',
            'class_name': 'border-amber-400 bg-amber-50/10 text-amber-900',
            'icon_class': 'bg-amber-100 text-amber-600',
            'url': reverse('coach:student_brans_create', kwargs={'student_id': student.id}) if coach_view else reverse('student:brans_create'),
        }]

    insights = []
    for key, label, _color, _klass, names in _BRANS_HUB_SUBJECTS:
        subject_entries = [e for e in entries if e.ders.name in names]
        recent_three = subject_entries[:3]
        if len(recent_three) == 3:
            newest, middle, oldest = recent_three
            drop = float(oldest.net) - float(newest.net)
            if newest.net < middle.net < oldest.net and drop >= 2:
                text = (
                    f"{student.full_name}'in {label} netleri düşüşte. Konulara bak →"
                    if coach_view else
                    f'{label} son sınavlarda düşüşte. Konulara bak →'
                )
                insights.append({
                    'text': text,
                    'severity': 'high',
                    'class_name': 'border-red-400 bg-red-50/10 text-red-900',
                    'icon_class': 'bg-red-100 text-red-600',
                    'url': _brans_insight_url(key, coach_view=coach_view, student_id=student.id),
                })

        period_entries = [e for e in subject_entries if e.tarih >= current_cutoff]
        previous_entries = [e for e in subject_entries if e.tarih < current_cutoff]
        if period_entries:
            period_best = max(period_entries, key=lambda e: e.net)
            previous_best = max((float(e.net) for e in previous_entries), default=None)
            if previous_best is None or float(period_best.net) > previous_best:
                text = (
                    f"{student.full_name} {label} branşında en iyi performansını yaptı!"
                    if coach_view else
                    f'{label} branşında en iyi performansın!'
                )
                insights.append({
                    'text': text,
                    'severity': 'low',
                    'class_name': 'border-emerald-400 bg-emerald-50/10 text-emerald-900',
                    'icon_class': 'bg-emerald-100 text-emerald-600',
                    'url': _brans_insight_url(key, coach_view=coach_view, student_id=student.id),
                })

    insights.sort(key=lambda item: {'high': 0, 'low': 1, 'empty': 2}.get(item['severity'], 3))
    return insights[:3]


def _build_brans_subject_detail_context(student, subject_slug, period='30g', *, coach_view=False):
    from itertools import groupby as _groupby
    from django.db.models import Sum

    subject_info = None
    for key, label, color, klass, names in _BRANS_HUB_SUBJECTS:
        if key == subject_slug:
            subject_info = {'key': key, 'name': label, 'color': color, 'class': klass, 'names': names}
            break
            
    if not subject_info:
        return None

    today = date.today()
    period_days = {'7g': 7, '30g': 30, '3 ay': 90, '1 yıl': 365}.get(period)
    current_cutoff = _brans_period_cutoff(period)
    previous_cutoff = today - timedelta(days=period_days * 2) if period_days else date.min

    # Fetch BransDeneme entries
    brans_qs = (
        BransDeneme.objects
        .filter(student=student, ders__name__in=subject_info['names'])
        .select_related('ders')
        .order_by('-tarih', '-id')
    )
    
    # Fetch ExamResult entries for Genel Denemeler
    genel_qs = (
        ExamResult.objects
        .filter(exam__student=student, subject__name__in=subject_info['names'])
        .select_related('exam', 'subject')
        .order_by('-exam__exam_date', '-exam__id')
    )
    
    # We will combine and annotate them
    exam_objs = []
    
    for b in brans_qs:
        exam_objs.append(SimpleNamespace(
            id=b.id,
            date=b.tarih,
            date_label=_brans_exam_label(b),
            month_label=f'{_TURKISH_MONTHS[b.tarih.month]} {b.tarih.year}',
            type='Branş',
            net=b.net,
            correct=b.dogru,
            wrong=b.yanlis,
            blank=b.bos,
            duration=f'{b.sure_dakika} dk' if b.sure_dakika else 'Süre yok',
            source='brans',
            compare_key=f'brans-{b.id}',
            detail_url=(
                reverse('coach:brans_student_detail', kwargs={'student_id': student.id})
                if coach_view else reverse('student:brans_detail', kwargs={'pk': b.id})
            ),
        ))
        
    for r in genel_qs:
        duration = f'{r.subject_duration_minutes} dk' if r.subject_duration_minutes else 'Süre yok'
        if not r.subject_duration_minutes and r.exam.duration_minutes:
             # fallback to exam duration if subject duration isn't set
             duration = f'{r.exam.duration_minutes} dk (Top)'
             
        exam_objs.append(SimpleNamespace(
            id=r.exam.id,
            date=r.exam.exam_date,
            date_label=f"{r.exam.exam_date.day} {_TURKISH_MONTHS[r.exam.exam_date.month]}",
            month_label=f'{_TURKISH_MONTHS[r.exam.exam_date.month]} {r.exam.exam_date.year}',
            type='Genel',
            net=float(r.net_score),
            correct=r.correct_answers,
            wrong=r.wrong_answers,
            blank=r.blank_answers,
            duration=duration,
            source='genel',
            compare_key=f'genel-{r.exam.id}',
            detail_url=(
                reverse('coach:student_detail', kwargs={'student_id': student.id})
                if coach_view else reverse('student:exam_detail_v2', kwargs={'exam_id': r.exam.id})
            ),
        ))
        
    # Sort by date descending
    exam_objs.sort(key=lambda e: e.date, reverse=True)
    
    stats_by_source = {
        'all': _subject_detail_source_stats(exam_objs, current_cutoff, previous_cutoff),
        'genel': _subject_detail_source_stats([e for e in exam_objs if e.source == 'genel'], current_cutoff, previous_cutoff),
        'brans': _subject_detail_source_stats([e for e in exam_objs if e.source == 'brans'], current_cutoff, previous_cutoff),
    }
    chart_points = _subject_detail_chart_points(exam_objs, current_cutoff)
    
    # Topic errors — BRANCH EXAMS ONLY (BransTopicError; trial exam errors excluded)
    brans_errors = (
        BransTopicError.objects
        .filter(brans_deneme__student=student, topic__subject__name__in=subject_info['names'])
    )
    if period_days:
        brans_errors = brans_errors.filter(brans_deneme__tarih__gte=current_cutoff)
    brans_errors = brans_errors.values('topic__id', 'topic__name').annotate(total_errors=Sum('yanlis_sayisi'))

    topic_errors_map = {}
    study_base_url = reverse('tasks:hafta') if coach_view else reverse('student_tasks:hafta')
    for r in brans_errors:
        tid = r['topic__id']
        if tid not in topic_errors_map:
            topic_errors_map[tid] = {'id': tid, 'name': r['topic__name'], 'errors': 0}
        topic_errors_map[tid]['errors'] += (r['total_errors'] or 0)

    for topic in topic_errors_map.values():
        query = {'topic_id': topic['id'], 'mode': 'konu_anlatimi'}
        if coach_view:
            query['student_id'] = student.id
        topic['study_url'] = f'{study_base_url}?{urlencode(query)}'
        
    topics = sorted(topic_errors_map.values(), key=lambda x: x['errors'], reverse=True)[:10]

    # ── Per-topic exam breakdown (accordion drill-down data) ─────────────────
    top_topic_ids = [t['id'] for t in topics]

    brans_by_topic = {}
    if top_topic_ids:
        bqs = (
            BransTopicError.objects
            .filter(topic__id__in=top_topic_ids, brans_deneme__student=student)
            .select_related('brans_deneme__ders', 'topic')
            .order_by('-brans_deneme__tarih')
        )
        if period_days:
            bqs = bqs.filter(brans_deneme__tarih__gte=current_cutoff)
        for err in bqs:
            tid = err.topic_id
            bd  = err.brans_deneme
            bid = bd.id
            brans_by_topic.setdefault(tid, {})
            if bid not in brans_by_topic[tid]:
                # coaches see the subject drill-down; students see per-entry brans_detail
                brans_by_topic[tid][bid] = {
                    'source': 'brans',
                    'id':     bid,
                    'name':   f'{bd.ders.name} — {_brans_exam_label(bd)}',
                    'date':   bd.tarih.isoformat(),
                    'errors': 0,
                    'url':    (
                        None  # no per-entry coach brans view; shown as info only
                        if coach_view
                        else reverse('student:brans_detail', kwargs={'pk': bid})
                    ),
                }
            brans_by_topic[tid][bid]['errors'] += err.yanlis_sayisi

    for topic in topics:
        tid = topic['id']
        exams = list(brans_by_topic.get(tid, {}).values())
        exams.sort(key=lambda x: x['date'], reverse=True)
        topic['exams'] = exams[:8]

    # ── Subject-specific branch task panel data (student & coach view) ──────
    branch_radar_json = '[]'
    branch_active_json = '[]'
    branch_completed_json = '[]'
    if True:  # always fetch; coach views the student's tasks
        assigned_topic_ids = set(
            StudentTask.objects.filter(
                student=student,
                topic__subject__name__in=subject_info['names'],
                task_source=StudentTask.SOURCE_BRANCH,
            ).values_list('topic_id', flat=True)
        )
        # Radar: all subject error-topics not yet assigned as BRANCH tasks
        all_subject_topics = sorted(topic_errors_map.values(), key=lambda x: x['errors'], reverse=True)
        branch_radar = [
            {'id': t['id'], 'name': t['name'], 'errors': t['errors'], 'sub': ''}
            for t in all_subject_topics
            if t['id'] not in assigned_topic_ids
        ][:20]

        branch_active_qs = (
            StudentTask.objects
            .filter(student=student, topic__subject__name__in=subject_info['names'],
                    is_completed=False, task_source=StudentTask.SOURCE_BRANCH)
            .select_related('topic')
            .order_by('-created_at')
        )
        branch_active = [
            {
                'id': t.id, 'topic_id': t.topic_id,
                'name': t.topic.name, 'sub': t.topic.sub_category or '',
                'byCoach': t.assigned_by_coach,
                'created_at': _fmt_task_date(t.created_at),
            }
            for t in branch_active_qs
        ]

        branch_completed_qs = (
            StudentTask.objects
            .filter(student=student, topic__subject__name__in=subject_info['names'],
                    is_completed=True, task_source=StudentTask.SOURCE_BRANCH)
            .select_related('topic')
            .order_by('-completed_at')[:30]
        )
        branch_completed = [
            {'id': t.id, 'topic_id': t.topic_id, 'name': t.topic.name, 'sub': t.topic.sub_category or '',
             'created_at':   _fmt_task_date(t.created_at),
             'completed_at': _fmt_task_date(t.completed_at),
             'cycle_days':   _task_cycle_days(t.created_at, t.completed_at)}
            for t in branch_completed_qs
        ]

        branch_radar_json = json.dumps(branch_radar)
        branch_active_json = json.dumps(branch_active)
        branch_completed_json = json.dumps(branch_completed)

    # Convert exam_objs to dict for easy JSON serialization if needed, or pass directly
    # The UI will filter by dataSource ('brans', 'genel'). We pass all.
    
    # History grouping by month
    cutoff = today - timedelta(days=30)
    recent = [e for e in exam_objs if e.date >= cutoff]
    older = [e for e in exam_objs if e.date < cutoff]
    archive = []
    for (year, month), group in _groupby(older, lambda e: (e.date.year, e.date.month)):
        month_exams = list(group)
        archive.append(SimpleNamespace(
            month_label=f'{_TURKISH_MONTHS[month]} {year}',
            count=len(month_exams),
            exams=month_exams,
        ))

    exams_json = json.dumps([
        {
            'id': e.id,
            'key': e.compare_key,
            'date': e.date_label,
            'isoDate': e.date.isoformat(),
            'monthLabel': 'Son 30 Gün' if e.date >= cutoff else e.month_label,
            'type': e.type,
            'net': float(e.net),
            'correct': e.correct,
            'wrong': e.wrong,
            'blank': e.blank,
            'duration': e.duration,
            'source': e.source,
            'url': e.detail_url,
        }
        for e in exam_objs
    ])

    return {
        'v2_shell': True,
        'shell_hide_fab': True,
        'shell_active': 'Branş',
        'student': student,
        'student_initials': _brans_student_initials(student),
        'coach_view': coach_view,
        'active_period': period,
        'period_choices': ['7g', '30g', '3 ay', '1 yıl', 'Tümü'],
        'subject': subject_info['name'],
        'subjectColor': subject_info['color'],
        'subjectClass': subject_info['class'],
        'subjectSlug': subject_slug,
        'export_period': period,
        'stats': stats_by_source['all'],
        'stats_by_source_json': json.dumps(stats_by_source),
        'chart_points_json': json.dumps(chart_points),
        'topics': topics,
        'exams': SimpleNamespace(
            recent=recent,
            archive=archive,
            total_count=len(exam_objs),
            all_json=exams_json
        ),
        'branch_radar_json':     branch_radar_json,
        'branch_active_json':    branch_active_json,
        'branch_completed_json': branch_completed_json,
    }

@student_required
def brans_subject_detail_student(request, subject_slug):
    period = request.GET.get('period', '30g')
    ctx = _build_brans_subject_detail_context(request.user, subject_slug, period)
    if not ctx:
        return redirect('brans_hub')
    return render(request, 'student/subject_drill_down.html', ctx)

@coach_required
def brans_subject_detail_coach(request, student_id, subject_slug):
    if not coach_can_view_student(request.user, student_id):
        return HttpResponseForbidden('Bu öğrenciyi görüntüleme yetkiniz yok.')
    from users_app.models import User as UserModel
    student = get_object_or_404(UserModel, id=student_id, role='student')
    period = request.GET.get('period', '30g')
    ctx = _build_brans_subject_detail_context(student, subject_slug, period, coach_view=True)
    if not ctx:
        return redirect('coach:student_brans_hub', student_id=student.id)
    return render(request, 'student/subject_drill_down.html', ctx)

def _build_brans_hub_context(student, period='30g', *, coach_view=False):
    """DAS-602/603/605: SSR data for the Branş Hub."""
    from itertools import groupby as _groupby

    today = date.today()
    period = _brans_export_period_query(period)
    period_days = {'7g': 7, '30g': 30, '3 ay': 90, '1 yıl': 365}.get(period)
    current_cutoff = _brans_period_cutoff(period)
    previous_cutoff = today - timedelta(days=period_days * 2) if period_days else date.min

    qs = (
        BransDeneme.objects
        .filter(student=student)  # type='branch' by model: only Branş Denemeleri live here.
        .select_related('ders')
        .order_by('-tarih', '-id')
    )
    entries = list(qs)
    insights = _build_brans_insights(student, entries, period, coach_view=coach_view)

    current = {key: [] for key, *_ in _BRANS_HUB_SUBJECTS}
    previous = {key: [] for key, *_ in _BRANS_HUB_SUBJECTS}
    for entry in entries:
        bucket = _brans_subject_bucket(entry.ders.name)
        if not bucket:
            continue
        if entry.tarih >= current_cutoff:
            current[bucket].append(entry.net)
        elif previous_cutoff <= entry.tarih < current_cutoff:
            previous[bucket].append(entry.net)

    subjects = []
    for key, label, color, klass, _names in _BRANS_HUB_SUBJECTS:
        cur_avg = round(sum(current[key]) / len(current[key]), 1) if current[key] else 0.0
        prev_avg = round(sum(previous[key]) / len(previous[key]), 1) if previous[key] else 0.0
        delta = round(cur_avg - prev_avg, 2) if current[key] and previous[key] else 0.0
        subjects.append(SimpleNamespace(
            key=key,
            name=label,
            net=cur_avg,
            delta=delta,
            trend='down' if delta < 0 else 'up',
            color=color,
            class_name=klass,
            count=len(current[key]),
        ))

    exam_objs = [
        SimpleNamespace(
            id=e.id,
            type='branch',
            date=e.tarih,
            date_label=_brans_exam_label(e),
            subject=e.ders.display_name,
            subject_key=_brans_subject_bucket(e.ders.name) or '',
            net=e.net,
            dogru=e.dogru,
            yanlis=e.yanlis,
            bos=e.bos,
            duration=f'{e.sure_dakika} dk' if e.sure_dakika else 'Süre yok',
            color=next((color for _k, _l, color, _c, names in _BRANS_HUB_SUBJECTS if e.ders.name in names), 'var(--primary)'),
            detail_url=reverse('student:brans_detail', kwargs={'pk': e.id}) if not coach_view else '#',
            edit_url=reverse('student:brans_edit', kwargs={'pk': e.id}) if not coach_view else '#',
            delete_url=reverse('student:brans_delete', kwargs={'pk': e.id}) if not coach_view else '#',
        )
        for e in entries
    ]

    cutoff = today - timedelta(days=30)
    recent = [e for e in exam_objs if e.date >= cutoff]
    older = [e for e in exam_objs if e.date < cutoff]
    archive = []
    for (year, month), group in _groupby(older, lambda e: (e.date.year, e.date.month)):
        month_exams = list(group)
        archive.append(SimpleNamespace(
            month_label=f'{_TURKISH_MONTHS[month]} {year}',
            count=len(month_exams),
            exams=month_exams,
        ))

    compare_entries_json = json.dumps([
        {
            'id': e.id,
            'subject': e.subject,
            'subjectKey': e.subject_key,
            'date': e.date_label,
            'monthLabel': 'Son 30 Gün' if e.date >= cutoff else f'{_TURKISH_MONTHS[e.date.month]} {e.date.year}',
            'dogru': e.dogru,
            'yanlis': e.yanlis,
            'bos': e.bos,
            'net': float(e.net),
            'duration': e.duration,
        }
        for e in exam_objs
    ])

    # ── Branch trend chart data (all-time; not period-filtered) ─────────────
    all_dates_sorted = sorted(set(e.tarih for e in entries))
    all_date_isos    = [d.isoformat() for d in all_dates_sorted]
    all_date_labels  = [f'{d.day} {_TURKISH_MONTHS[d.month]}' for d in all_dates_sorted]

    chart_datasets = []
    for key, label, color, _klass, names in _BRANS_HUB_SUBJECTS:
        subj_map = {
            e.tarih.isoformat(): round(float(e.net), 2)
            for e in entries if e.ders.name in names
        }
        if subj_map:
            chart_datasets.append({
                'label':              label,
                'data':               [subj_map.get(iso, None) for iso in all_date_isos],
                'borderColor':        color,
                'backgroundColor':    'transparent',
                'borderWidth':        2.5,
                'pointBackgroundColor': color,
                'pointRadius':        4,
                'pointHoverRadius':   7,
                'tension':            0.35,
                'fill':               False,
                'spanGaps':           True,
            })

    brans_chart_json = json.dumps({
        'labels':   all_date_labels,
        'datasets': chart_datasets,
        'has_data': bool(all_date_labels),
    })

    return {
        'v2_shell': True,
        'shell_hide_fab': True,
        'shell_active': 'Branş',
        'student': student,
        'student_initials': _brans_student_initials(student),
        'coach_view': coach_view,
        'brans_add_url': f"/coach/student/{student.id}/brans/ekle/?next=/coach/student/{student.id}/brans/" if coach_view else '/student/brans/ekle/?next=/brans/',
        'brans_add_label': f'{student.full_name} için Branş Denemesi Ekle' if coach_view else 'Branş Denemesi Ekle',
        'active_period': period,
        'period_choices': ['7g', '30g', '3 ay', '1 yıl', 'Tümü'],
        'insights': insights,
        'export_period': period,
        'compare_entries_json': compare_entries_json,
        'subjects': subjects,
        'brans_chart_json': brans_chart_json,
        'exams': SimpleNamespace(
            recent=recent,
            archive=archive,
            total_count=len(exam_objs),
        ),
    }


@student_required
def brans_hub_student(request):
    period = request.GET.get('period', '30g')
    return render(request, 'student/brans_hub.html', _build_brans_hub_context(request.user, period))


@coach_required
def brans_hub_coach(request, student_id):
    if not coach_can_view_student(request.user, student_id):
        return HttpResponseForbidden('Bu öğrenciyi görüntüleme yetkiniz yok.')
    from users_app.models import User as UserModel
    student = get_object_or_404(UserModel, id=student_id, role='student')
    period = request.GET.get('period', '30g')
    return render(request, 'student/brans_hub.html', _build_brans_hub_context(student, period, coach_view=True))


def _brans_export_student(request):
    if not request.user.is_authenticated:
        return None, HttpResponseForbidden('Giriş gerekli.')
    student_id = request.GET.get('student_id')
    if request.user.role == 'coach':
        if not student_id or not coach_can_view_student(request.user, student_id):
            return None, HttpResponseForbidden('Bu öğrenciyi dışa aktarma yetkiniz yok.')
        from users_app.models import User as UserModel
        return get_object_or_404(UserModel, id=student_id, role='student'), None
    if student_id and str(request.user.id) != str(student_id):
        return None, HttpResponseForbidden('Bu veriyi dışa aktarma yetkiniz yok.')
    return request.user, None


def _brans_export_entries(student, period, subject_slug=None):
    cutoff = _brans_period_cutoff(period)
    qs = BransDeneme.objects.filter(student=student).select_related('ders').order_by('-tarih', '-id')
    if cutoff != date.min:
        qs = qs.filter(tarih__gte=cutoff)
    if subject_slug:
        subject_names = next((names for key, _label, _color, _klass, names in _BRANS_HUB_SUBJECTS if key == subject_slug), None)
        if subject_names:
            qs = qs.filter(ders__name__in=subject_names)
    return list(qs)


def brans_export_xlsx(request):
    student, error_response = _brans_export_student(request)
    if error_response:
        return error_response

    period = _brans_export_period_query(request.GET.get('period', '30g'))
    entries = _brans_export_entries(student, period, request.GET.get('subject'))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Brans Hub'
    headers = ['Tarih', 'Ders', 'Dogru', 'Yanlis', 'Bos', 'Net', 'Sure (dk)', 'Not']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='FFEDE9FE')
    for entry in entries:
        ws.append([
            entry.tarih.isoformat(),
            entry.ders.display_name,
            entry.dogru,
            entry.yanlis,
            entry.bos,
            entry.net,
            entry.sure_dakika or '',
            entry.ogrenci_notu or '',
        ])
    for column in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 40)

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="brans_{student.id}_{period}.xlsx"'
    return response


def brans_export_html(request):
    student, error_response = _brans_export_student(request)
    if error_response:
        return error_response

    period = _brans_export_period_query(request.GET.get('period', '30g'))
    entries = _brans_export_entries(student, period, request.GET.get('subject'))

    from django.utils.html import escape

    rows = ''.join(
        '<tr>'
        f'<td>{escape(entry.tarih.isoformat())}</td>'
        f'<td>{escape(entry.ders.display_name)}</td>'
        f'<td>{entry.dogru}</td>'
        f'<td>{entry.yanlis}</td>'
        f'<td>{entry.bos}</td>'
        f'<td>{entry.net:.2f}</td>'
        f'<td>{entry.sure_dakika or ""}</td>'
        f'<td>{escape(entry.ogrenci_notu or "")}</td>'
        '</tr>'
        for entry in entries
    )
    html = f'''<!doctype html>
<html lang="tr"><head><meta charset="utf-8"><title>Branş Hub Export</title>
<style>body{{font-family:Arial,sans-serif;padding:24px}}table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #ddd;padding:8px;text-align:left}}th{{background:#ede9fe}}</style>
</head><body><h1>{escape(student.full_name)} - Branş Hub</h1><p>Periyot: {escape(period)}</p>
<table><thead><tr><th>Tarih</th><th>Ders</th><th>Doğru</th><th>Yanlış</th><th>Boş</th><th>Net</th><th>Süre</th><th>Not</th></tr></thead><tbody>{rows}</tbody></table>
</body></html>'''
    response = HttpResponse(html, content_type='text/html; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="brans_{student.id}_{period}.html"'
    return response


@coach_required
def coach_brans_create_for_student(request, student_id):
    if not coach_can_view_student(request.user, student_id):
        return HttpResponseForbidden('Bu öğrenciyi görüntüleme yetkiniz yok.')
    from users_app.models import User as UserModel
    student = get_object_or_404(UserModel, id=student_id, role='student')
    next_url = _safe_next_url(request, f'/coach/student/{student.id}/brans/')

    if request.method == 'POST':
        form = BransDenemeForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.student = student
            entry.save()
            topic_url = reverse('coach:student_brans_topic_errors', kwargs={'student_id': student.id, 'pk': entry.pk})
            return redirect(_append_next(topic_url, next_url))
    else:
        form = BransDenemeForm()

    subjects_json = json.dumps({str(s.pk): s.question_count for s in Subject.objects.only('pk', 'question_count')})
    return render(request, 'student/brans_create.html', {
        'form': form,
        'subjects_json': subjects_json,
        'next_url': next_url,
        'back_url': next_url,
        'coach_view': True,
        'student': student,
        'page_title': f'{student.full_name} için Branş Denemesi Ekle',
        'submit_label': 'Branş Denemesini Kaydet',
    })


@student_required
def dashboard_analytics_partial(request):
    """AJAX endpoint for period-filter updates — returns JSON with HTML fragments + chart data.
    Avoids full-page reload while keeping all analytics sections semantically consistent."""
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return redirect('student:dashboard')

    from django.template.loader import render_to_string

    period, since = _period_since(request.GET.get('period', 'all'))
    chart_period, chart_since, chart_period_label = _chart_period_since(request.GET.get('chart_period', '30'))

    exams_qs = (
        Exam.objects.filter(student=request.user)
        .prefetch_related('results__subject')
        .order_by('-exam_date')
    )
    if since:
        exams_qs = exams_qs.filter(exam_date__gte=since)
    exams = exams_qs[:20]

    chart_data   = _build_student_chart_data(request.user.id, since=chart_since)
    scatter_data = _build_subject_duration_scatter_data(request.user.id, since=chart_since)
    radar_topics = _get_radar_topics(request.user.id, since=since)

    exam_list = list(exams)
    v2_stats  = _build_v2_stats(exam_list, period)
    v2_stats['radar_count'] = len(radar_topics)
    subjects   = _build_v2_subjects(exam_list)

    ctx = {
        'stats':       v2_stats,
        'subjects':    subjects,
        'exams':       exams,
        'period':      period,
        'chart_period': chart_period,
        'chart_period_label': chart_period_label,
    }

    def _render(tpl, extra=None):
        c = dict(ctx, **(extra or {}))
        return render_to_string(tpl, c, request=request)

    # radar_data feeds the task panel's wrong-topic list (period-filtered)
    radar_data = [
        {'id': t['topic__id'], 'name': t['topic__name'],
         'subject': t['topic__subject__name'],
         'sub': t['topic__sub_category'] or '',
         'wrong': t['wrong_total'] or 0, 'blank': t['blank_total'] or 0}
        for t in radar_topics
    ]

    return JsonResponse({
        'greeting_html': _render('dashboard/partials/_greeting.html'),
        'kpi_html':      _render('dashboard/partials/_kpi_strip.html'),
        'subjects_html': _render('dashboard/partials/_subject_breakdown_carousel.html') if subjects else '',
        'chart_data':    chart_data,
        'scatter_data':  scatter_data,
        'has_exams':     bool(chart_data.get('labels')),
        'radar_data':    radar_data,
        'period':        period,
        'chart_period':  chart_period,
    })


@student_required
def denemeler_list(request):
    # DAS-302: flag gate — DB field wins; query param overrides for dev/testing
    is_v2 = (request.user.denemeler_v2
             or request.GET.get('denemeler_v2') == '1'
             or getattr(settings, 'V2_DEFAULT', False))
    if is_v2:
        filters = {
            'period':   request.GET.get('period', ''),
            'sort':     request.GET.get('sort', 'date'),
            'has_note': request.GET.get('has_note', '') == '1',
        }
        ctx = _build_denemeler_v2_context(request.user, filters=filters)
        ctx['filters'] = filters
        ctx['trial_topics'] = _build_trial_topic_errors(request.user, filters['period'])
        return render(request, 'student/denemeler_v2.html', ctx)
    # V1 path — unchanged
    from django.core.paginator import Paginator
    qs = (
        Exam.objects
        .filter(student=request.user)
        .select_related('publisher')
        .order_by('-exam_date', '-id')
    )
    page = Paginator(qs, 20).get_page(request.GET.get('p'))
    return render(request, 'student/denemeler.html', {'page': page})


def _period_since(period_str):
    """Return (period_label, since_date) from a ?period= query-string value."""
    if period_str == '30':
        return period_str, date.today() - timedelta(days=30)
    if period_str == '90':
        return period_str, date.today() - timedelta(days=90)
    return 'all', None


def _build_trial_topic_errors(user, period_str='', top_n=12, *, coach_view=False, student_id=None):
    """Top error topics from TRIAL exams, with per-exam breakdown, for the accordion.

    Pass coach_view=True and student_id=<int> so drill-down links point to the coach
    exam detail route instead of the student route.
    """
    from django.db.models import Sum as _Sum

    qs = ExamTopicError.objects.filter(exam__student=user)
    since = None
    if period_str in _PERIOD_DAYS:
        since = date.today() - timedelta(days=_PERIOD_DAYS[period_str])
        qs = qs.filter(exam__exam_date__gte=since)

    agg = list(
        qs.values('topic__id', 'topic__name', 'topic__subject__name')
        .annotate(total=_Sum('wrong_count') + _Sum('blank_count'))
        .order_by('-total')[:top_n]
    )
    if not agg:
        return []

    top_topic_ids = [r['topic__id'] for r in agg]
    detail_qs = (
        ExamTopicError.objects
        .filter(topic__id__in=top_topic_ids, exam__student=user)
        .select_related('exam__publisher', 'topic')
        .order_by('-exam__exam_date')
    )
    if since is not None:
        detail_qs = detail_qs.filter(exam__exam_date__gte=since)

    exam_by_topic = {}
    for err in detail_qs:
        tid = err.topic_id
        eid = err.exam_id
        exam_by_topic.setdefault(tid, {})
        if eid not in exam_by_topic[tid]:
            if coach_view and student_id:
                url = reverse('coach:student_exam_detail',
                              kwargs={'student_id': student_id, 'exam_id': eid})
            else:
                url = reverse('student:exam_detail_v2', kwargs={'exam_id': eid})
            exam_by_topic[tid][eid] = {
                'id':     eid,
                'name':   err.exam.custom_name,
                'date':   err.exam.exam_date.isoformat(),
                'errors': 0,
                'url':    url,
            }
        exam_by_topic[tid][eid]['errors'] += (err.wrong_count or 0) + (err.blank_count or 0)

    topics = []
    for r in agg:
        tid = r['topic__id']
        exams = sorted(exam_by_topic.get(tid, {}).values(), key=lambda x: x['date'], reverse=True)[:8]
        topics.append({
            'id':      tid,
            'name':    r['topic__name'],
            'subject': r['topic__subject__name'] or '',
            'errors':  r['total'] or 0,
            'exams':   exams,
        })
    return topics


def _chart_period_since(period_str):
    if period_str == '7':
        return '7', date.today() - timedelta(days=7), 'Son 7 gün'
    if period_str == '30':
        return '30', date.today() - timedelta(days=30), 'Son 30 gün'
    return 'all', None, 'Tümü'


@student_required
def student_dashboard(request):
    period, since = _period_since(request.GET.get('period', 'all'))
    chart_period, chart_since, chart_period_label = _chart_period_since(request.GET.get('chart_period', '30'))

    exams_qs = (
        Exam.objects.filter(student=request.user)
        .prefetch_related('results__subject')
        .order_by('-exam_date')
    )
    if since:
        exams_qs = exams_qs.filter(exam_date__gte=since)
    exams = exams_qs[:20]

    chart_data = _build_student_chart_data(request.user.id, since=chart_since)
    scatter_data = _build_subject_duration_scatter_data(request.user.id, since=chart_since)
    radar_topics = _get_radar_topics(request.user.id, since=since)

    active_qs = (
        StudentTask.objects.filter(student=request.user, is_completed=False,
                                   task_source=StudentTask.SOURCE_TRIAL)
        .select_related('topic__subject').order_by('created_at')
    )
    completed_qs = (
        StudentTask.objects.filter(student=request.user, is_completed=True,
                                   task_source=StudentTask.SOURCE_TRIAL)
        .select_related('topic__subject')
        .order_by('topic__subject__name', '-completed_at')[:50]
    )

    radar_json = json.dumps([
        {'id': t['topic__id'], 'name': t['topic__name'], 'subject': t['topic__subject__name'],
         'sub': t['topic__sub_category'] or '', 'wrong': t['wrong_total'] or 0, 'blank': t['blank_total'] or 0}
        for t in radar_topics
    ])
    active_json = json.dumps([
        {'id': t.id, 'topic_id': t.topic_id, 'name': t.topic.name, 'subject': t.topic.subject.name,
         'sub': t.topic.sub_category or '', 'byCoach': t.assigned_by_coach,
         'created_at': _fmt_task_date(t.created_at)}
        for t in active_qs
    ])
    completed_json = json.dumps([
        {'id': t.id, 'topic_id': t.topic_id, 'name': t.topic.name, 'subject': t.topic.subject.name,
         'sub': t.topic.sub_category or '',
         'created_at':  _fmt_task_date(t.created_at),
         'completed_at': _fmt_task_date(t.completed_at),
         'cycle_days':  _task_cycle_days(t.created_at, t.completed_at)}
        for t in completed_qs
    ])

    dashboard_v2 = request.GET.get('v2') == '1' or getattr(settings, 'V2_DEFAULT', False)
    exam_list = list(exams)

    ctx = {
        'exams': exams,
        'chart_data_json': json.dumps(chart_data),
        'scatter_data_json': json.dumps(scatter_data),
        'has_chart_data': bool(chart_data.get('labels')),
        'radar_json': radar_json,
        'active_json': active_json,
        'completed_json': completed_json,
        'period': period,
        'chart_period': chart_period,
        'chart_period_label': chart_period_label,
        'dashboard_v2': dashboard_v2,
    }

    if dashboard_v2:
        v2_stats = _build_v2_stats(exam_list, period)
        v2_stats['radar_count'] = len(radar_topics)
        ctx['stats'] = v2_stats
        ctx['subjects'] = _build_v2_subjects(exam_list)

    ctx['first_run'] = request.session.pop('first_run', False)
    template = 'student/dashboard_v2.html' if dashboard_v2 else 'student/dashboard.html'
    return render(request, template, ctx)


# ──────────────────────────────────────────────
# STUDENT — EXAM CRUD
# ──────────────────────────────────────────────

_SUBJECT_COLORS_V2 = {
    'TYT Türkçe':          '#3B82F6',
    'TYT Matematik':       '#A78BFA',
    'TYT Sosyal Bilimler': '#F59E0B',
    'TYT Fen Bilimleri':   '#10B981',
}


def _build_exam_detail_v2_context(exam, results_qs, topic_errors_qs):
    """DAS-321–325: rich context for the V2 exam detail page."""
    from itertools import groupby as _groupby

    results = list(results_qs)
    te_list  = list(topic_errors_qs)

    # ── Subject breakdown ─────────────────────────────────────────────────
    subject_breakdown = []
    for r in results:
        total_q = r.correct_answers + r.wrong_answers + r.blank_answers
        accuracy = round(r.correct_answers / total_q * 100) if total_q > 0 else 0
        subject_breakdown.append(SimpleNamespace(
            id=r.id,
            name=r.subject.display_name,
            full_name=r.subject.name,
            color=_SUBJECT_COLORS_V2.get(r.subject.name, '#6D5BFF'),
            net=round(r.net_score, 2),
            accuracy_percent=accuracy,
            correct=r.correct_answers,
            wrong=r.wrong_answers,
            blank=r.blank_answers,
            question_count=r.subject.question_count,
        ))

    net_total = round(sum(r.net_score for r in results), 2)

    # ── Previous exam (same student, immediately before this one) ─────────
    prev_exam_obj = (
        Exam.objects
        .filter(
            student_id=exam.student_id,
        )
        .filter(
            models.Q(exam_date__lt=exam.exam_date) |
            models.Q(exam_date=exam.exam_date, id__lt=exam.id)
        )
        .select_related('publisher')
        .order_by('-exam_date', '-id')
        .first()
    )

    net_delta         = None
    net_delta_display = None
    prev_exam_ctx     = None
    subject_deltas    = []

    if prev_exam_obj:
        prev_results = {
            r.subject.name: r
            for r in prev_exam_obj.results.select_related('subject').all()
        }
        prev_net = round(sum(r.net_score for r in prev_results.values()), 2)
        net_delta = round(net_total - prev_net, 2)
        from django.utils.formats import number_format as _nf
        _abs = _nf(abs(net_delta), decimal_pos=2)
        net_delta_display = f'+{_abs}' if net_delta >= 0 else f'-{_abs}'

        prev_exam_ctx = SimpleNamespace(
            id=prev_exam_obj.id,
            publisher=prev_exam_obj.publisher.name,
            date=prev_exam_obj.exam_date,
            net_total=prev_net,
        )

        # Per-subject deltas for DAS-327 compare sheet
        all_deltas = []
        for r in results:
            prev_r = prev_results.get(r.subject.name)
            delta = round(r.net_score - (prev_r.net_score if prev_r else 0), 2)
            all_deltas.append(SimpleNamespace(name=r.subject.display_name, delta=delta))
        max_abs = max((abs(s.delta) for s in all_deltas), default=1) or 1
        for s in all_deltas:
            s.delta_percent     = round(abs(s.delta) / max_abs * 100, 1) if s.delta > 0 else 0
            s.delta_percent_abs = round(abs(s.delta) / max_abs * 100, 1) if s.delta < 0 else 0
        subject_deltas = all_deltas

    # ── Topic error groups ────────────────────────────────────────────────
    has_topic_errors    = bool(te_list)
    total_topic_errors  = sum(te.wrong_count + te.blank_count for te in te_list)

    topic_error_groups = []
    for subj_name, group in _groupby(te_list, lambda te: te.topic.subject.name):
        group_list = list(group)
        topics = [
            SimpleNamespace(
                name=te.topic.name,
                sub_category=te.topic.sub_category,
                error_count=te.wrong_count + te.blank_count,
            )
            for te in group_list
            if (te.wrong_count + te.blank_count) > 0
        ]
        if topics:
            topic_error_groups.append(SimpleNamespace(
                name=group_list[0].topic.subject.display_name,
                color=_SUBJECT_COLORS_V2.get(subj_name, '#6D5BFF'),
                topics=topics,
            ))

    return {
        'exam':               exam,
        'net_total':          net_total,
        'net_delta':          net_delta,
        'net_delta_display':  net_delta_display,
        'subject_breakdown':  subject_breakdown,
        'has_topic_errors':   has_topic_errors,
        'total_topic_errors': total_topic_errors,
        'topic_error_groups': topic_error_groups,
        'prev_exam':          prev_exam_ctx,
        'subject_deltas':     subject_deltas,
        'exam_type':          'TYT',
    }


@student_required
def student_exam_detail(request, exam_id):
    """DAS-372: V1-only — redirects to V2 when denemeler_v2 flag is on."""
    is_v2 = (
        getattr(request.user, 'denemeler_v2', False)
        or request.GET.get('denemeler_v2') == '1'
        or getattr(settings, 'V2_DEFAULT', False)
    )
    if is_v2:
        return redirect('student:exam_detail_v2', exam_id=exam_id)

    exam = get_object_or_404(Exam, id=exam_id, student=request.user)
    results_qs = exam.results.select_related('subject').order_by('subject__name')
    topic_errors_qs = exam.topic_errors.select_related('topic__subject').order_by(
        'topic__subject__name', 'topic__sub_category', 'topic__name'
    )
    return render(request, 'student/exam_detail.html', {
        'exam': exam,
        'results': results_qs,
        'topic_errors': topic_errors_qs,
    })


@student_required
def student_exam_detail_v2(request, exam_id):
    """DAS-372: V2-only exam detail — redirects to V1 when flag is off."""
    is_v2 = (
        getattr(request.user, 'denemeler_v2', False)
        or request.GET.get('denemeler_v2') == '1'
        or getattr(settings, 'V2_DEFAULT', False)
    )
    if not is_v2:
        return redirect('student:exam_detail', exam_id=exam_id)

    exam = get_object_or_404(Exam, id=exam_id, student=request.user)
    results_qs = exam.results.select_related('subject').order_by('subject__name')
    topic_errors_qs = exam.topic_errors.select_related('topic__subject').order_by(
        'topic__subject__name', 'topic__sub_category', 'topic__name'
    )
    ctx = _build_exam_detail_v2_context(exam, results_qs, topic_errors_qs)
    return render(request, 'student/exam_detail_v2.html', ctx)


@student_required
@require_http_methods(['PATCH'])
def exam_result_update(request, exam_id, result_id):
    """DAS-326: PATCH — update a single subject's D/Y/B counts and recompute net scores."""
    exam   = get_object_or_404(Exam, id=exam_id, student=request.user)
    result = get_object_or_404(ExamResult, id=result_id, exam=exam)
    try:
        data    = json.loads(request.body)
        correct = max(0, int(data.get('correct', 0)))
        wrong   = max(0, int(data.get('wrong',   0)))
        blank   = max(0, int(data.get('blank',   0)))

        max_q = result.subject.question_count
        if correct + wrong + blank > max_q:
            return JsonResponse(
                {'status': 'error',
                 'detail': f'Toplam {max_q} soruyu geçemez ({result.subject.display_name}).'},
                status=400,
            )

        result.correct_answers = correct
        result.wrong_answers   = wrong
        result.blank_answers   = blank
        result.save()   # net_score = round(correct - wrong * 0.25, 2) in ExamResult.save()

        total_net = round(sum(r.net_score for r in exam.results.all()), 2)
        return JsonResponse({
            'status':    'ok',
            'net_score': result.net_score,
            'total_net': total_net,
        })
    except (json.JSONDecodeError, ValueError, TypeError):
        return JsonResponse({'status': 'error', 'detail': 'Geçersiz veri.'}, status=400)


@student_required
@require_http_methods(['PATCH'])
def exam_notes_update(request, exam_id):
    """DAS-325: PATCH endpoint for auto-saving exam notes."""
    exam = get_object_or_404(Exam, id=exam_id, student=request.user)
    try:
        data = json.loads(request.body)
        notes = data.get('notes', '') or ''
        exam.student_note = notes if notes.strip() else None
        exam.save(update_fields=['student_note'])
        return JsonResponse({'status': 'ok'})
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'status': 'error', 'detail': 'invalid json'}, status=400)


@student_required
def exam_edit(request, exam_id):
    """Edit an existing exam — reuses the V2 create template for a unified UI."""
    exam = get_object_or_404(Exam, id=exam_id, student=request.user)
    subjects   = _build_v2_entry_subjects()
    publishers = Publisher.objects.all()

    if request.method == 'POST':
        publisher_id = request.POST.get('publisher')
        custom_name  = request.POST.get('custom_name', '').strip()
        exam_date    = request.POST.get('exam_date')
        duration     = request.POST.get('duration_minutes') or None
        student_note = request.POST.get('student_note', '').strip() or None

        if not all([publisher_id, custom_name, exam_date]):
            messages.error(request, 'Lütfen zorunlu alanları doldurun.')
            return redirect('student:exam_edit', exam_id=exam.id)

        exam.publisher_id    = publisher_id
        exam.custom_name     = custom_name
        exam.exam_date       = exam_date
        exam.duration_minutes = duration
        exam.student_note    = student_note
        exam.save()

        for subject in subjects:
            correct = int(request.POST.get(f'correct_{subject.id}', 0) or 0)
            wrong   = int(request.POST.get(f'wrong_{subject.id}',   0) or 0)
            blank   = int(request.POST.get(f'blank_{subject.id}',   0) or 0)
            ExamResult.objects.update_or_create(
                exam=exam, subject_id=subject.id,
                defaults={'correct_answers': correct, 'wrong_answers': wrong, 'blank_answers': blank},
            )

        # Inline topic errors — replace all
        errors_raw = request.POST.get('errors_json', '[]')
        try:
            error_rows = json.loads(errors_raw)
        except (json.JSONDecodeError, ValueError):
            error_rows = []
        exam.topic_errors.all().delete()
        for row in error_rows:
            topic_id  = row.get('topicId')
            wrong_cnt = max(0, int(row.get('wrongCount', 0) or 0))
            blank_cnt = max(0, int(row.get('blankCount', 0) or 0))
            if topic_id and (wrong_cnt > 0 or blank_cnt > 0):
                ExamTopicError.objects.update_or_create(
                    exam=exam, topic_id=topic_id,
                    defaults={'wrong_count': wrong_cnt, 'blank_count': blank_cnt},
                )

        messages.success(request, 'Sınav güncellendi.')
        return redirect('student:exam_detail_v2', exam_id=exam.id)

    # Build prefill JSON for Alpine state initialisation
    result_map = {str(r.subject_id): r for r in exam.results.all()}
    rows_prefill = {
        str(s.id): {
            'd': result_map[str(s.id)].correct_answers if str(s.id) in result_map else 0,
            'y': result_map[str(s.id)].wrong_answers   if str(s.id) in result_map else 0,
            'total': s.question_count,
        }
        for s in subjects
    }
    topic_errors_prefill = [
        {
            'topicId':     te.topic_id,
            'topicName':   te.topic.name,
            'subjectId':   str(te.topic.subject_id),
            'subjectName': te.topic.subject.name,
            'wrongCount':  te.wrong_count,
            'blankCount':  te.blank_count,
        }
        for te in exam.topic_errors.select_related('topic__subject').all()
    ]
    prefill = {
        'publisherId': str(exam.publisher_id or ''),
        'customName':  exam.custom_name,
        'examDate':    exam.exam_date.isoformat(),
        'duration':    str(exam.duration_minutes or ''),
        'studentNote': exam.student_note or '',
        'rows':        rows_prefill,
        'topicRows':   topic_errors_prefill,
    }

    ctx = _v2_create_ctx(subjects, publishers, request)
    ctx.update({
        'is_edit':    True,
        'exam':       exam,
        'prefill_json': json.dumps(prefill),
    })
    return render(request, 'student/exam_create_v2.html', ctx)


@student_required
def exam_delete(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, student=request.user)
    if request.method == 'POST':
        exam.delete()
        messages.success(request, f'"{exam.custom_name}" silindi.')
        return redirect('student:dashboard')
    return render(request, 'student/exam_delete.html', {'exam': exam})


# ──────────────────────────────────────────────
# STUDENT — EXAM CREATE V2 (DAS-343 / Phase 3.3)
# Feature-flagged fast-entry; falls back to step1 when flag is off.
# ──────────────────────────────────────────────

_ENTRY_SUBJECT_COLORS = {
    'TYT Türkçe':          '#3B82F6',
    'TYT Matematik':       '#A78BFA',
    'TYT Sosyal Bilimler': '#F59E0B',
    'TYT Fen Bilimleri':   '#10B981',
    'TYT Fizik':           '#10B981',
    'TYT Kimya':           '#06B6D4',
    'TYT Biyoloji':        '#22C55E',
}


def _build_v2_entry_subjects():
    """Return SimpleNamespace list with color annotations for create v2."""
    qs = (
        Subject.objects.filter(exam_type='TYT')
        .exclude(name='TYT Fen Bilimleri')
        .order_by('name')
    )
    return [
        SimpleNamespace(
            id=s.id,
            name=s.name,
            display_name=s.display_name,
            question_count=s.question_count,
            color=_ENTRY_SUBJECT_COLORS.get(s.name, '#6D5BFF'),
        )
        for s in qs
    ]


def _build_topics_json():
    """Shared topics JSON for Alpine topic picker."""
    qs = Topic.objects.select_related('subject').order_by(
        'subject__name', 'sub_category', 'name'
    )
    return json.dumps([
        {
            'id': t.id,
            'name': t.name,
            'subject_id': t.subject_id,
            'subject_name': t.subject.name,
            'sub_category': t.sub_category,
        }
        for t in qs
    ])


@student_required
def exam_create_v2(request):
    """DAS-343 / DAS-372: V2-only fast-entry create. Redirects to V1 when flag is off."""
    is_v2 = (
        getattr(request.user, 'denemeler_v2', False)
        or request.GET.get('denemeler_v2') == '1'
        or getattr(settings, 'V2_DEFAULT', False)
    )
    if not is_v2:
        return redirect('student:exam_step1')

    subjects   = _build_v2_entry_subjects()
    publishers = Publisher.objects.all()

    if request.method == 'POST':
        publisher_id = request.POST.get('publisher')
        custom_name  = request.POST.get('custom_name', '').strip()
        exam_date    = request.POST.get('exam_date')
        duration     = request.POST.get('duration_minutes') or None
        student_note = request.POST.get('student_note', '').strip() or None

        if not all([publisher_id, custom_name, exam_date]):
            messages.error(request, 'Lütfen zorunlu alanları doldurun.')
            return render(request, 'student/exam_create_v2.html', _v2_create_ctx(subjects, publishers, request))

        exam = Exam.objects.create(
            student=request.user,
            publisher_id=publisher_id,
            custom_name=custom_name,
            exam_date=exam_date,
            duration_minutes=duration,
            student_note=student_note,
        )

        for subject in subjects:
            correct = int(request.POST.get(f'correct_{subject.id}', 0) or 0)
            wrong   = int(request.POST.get(f'wrong_{subject.id}',   0) or 0)
            blank   = int(request.POST.get(f'blank_{subject.id}',   0) or 0)
            ExamResult.objects.create(
                exam=exam, subject_id=subject.id,
                correct_answers=correct, wrong_answers=wrong, blank_answers=blank,
            )

        # Inline topic errors (DAS-348)
        errors_raw = request.POST.get('errors_json', '[]')
        try:
            error_rows = json.loads(errors_raw)
        except (json.JSONDecodeError, ValueError):
            error_rows = []
        for row in error_rows:
            topic_id   = row.get('topicId')
            wrong_cnt  = max(0, int(row.get('wrongCount', 0) or 0))
            blank_cnt  = max(0, int(row.get('blankCount', 0) or 0))
            if topic_id and (wrong_cnt > 0 or blank_cnt > 0):
                ExamTopicError.objects.update_or_create(
                    exam=exam, topic_id=topic_id,
                    defaults={'wrong_count': wrong_cnt, 'blank_count': blank_cnt},
                )

        from users_app.services.streak_engine import record_activity
        record_activity(request.user)
        messages.success(request, 'Sınav kaydedildi!')
        return redirect('student:exam_detail', exam_id=exam.id)

    return render(request, 'student/exam_create_v2.html', _v2_create_ctx(subjects, publishers, request))


def _v2_create_ctx(subjects, publishers, request):
    topics_json   = _build_topics_json()
    subjects_json = json.dumps([{'id': s.id, 'name': s.name} for s in subjects])
    return {
        'subjects':      subjects,
        'publishers':    publishers,
        'topics_json':   topics_json,
        'subjects_json': subjects_json,
        'today_iso':     date.today().isoformat(),
        'v2_shell':      True,
    }


# ──────────────────────────────────────────────
# STUDENT — EXAM STEP 1 (create)
# ──────────────────────────────────────────────

@student_required
def exam_create_step1(request):
    """DAS-372: V1 create — redirects to V2 when denemeler_v2 flag is on."""
    is_v2 = (
        getattr(request.user, 'denemeler_v2', False)
        or request.GET.get('denemeler_v2') == '1'
        or getattr(settings, 'V2_DEFAULT', False)
    )
    if is_v2:
        return redirect('student:exam_create_v2')

    subjects = Subject.objects.filter(exam_type='TYT').exclude(name='TYT Fen Bilimleri').order_by('name')
    publishers = Publisher.objects.all()

    if request.method == 'POST':
        publisher_id = request.POST.get('publisher')
        custom_name = request.POST.get('custom_name', '').strip()
        exam_date = request.POST.get('exam_date')
        duration = request.POST.get('duration_minutes') or None

        if not all([publisher_id, custom_name, exam_date]):
            messages.error(request, 'Lütfen zorunlu alanları doldurun.')
            return render(request, 'student/exam_create_step1.html', {'subjects': subjects, 'publishers': publishers})

        student_note = request.POST.get('student_note', '').strip() or None

        exam = Exam.objects.create(
            student=request.user,
            publisher_id=publisher_id,
            custom_name=custom_name,
            exam_date=exam_date,
            duration_minutes=duration,
            student_note=student_note,
        )

        total_duration = sum(
            int(request.POST.get(f'duration_{s.id}', '0') or 0)
            for s in subjects
        )
        if total_duration > 165:
            messages.error(request, 'Ders sürelerinin toplamı 165 dakikayı aşamaz.')
            return render(request, 'student/exam_create_step1.html',
                          {'subjects': subjects, 'publishers': publishers})

        has_errors = False
        for subject in subjects:
            correct = int(request.POST.get(f'correct_{subject.id}', 0) or 0)
            wrong = int(request.POST.get(f'wrong_{subject.id}', 0) or 0)
            blank = int(request.POST.get(f'blank_{subject.id}', 0) or 0)
            duration_raw = request.POST.get(f'duration_{subject.id}', '').strip()
            duration_val = int(duration_raw) if duration_raw.isdigit() and int(duration_raw) > 0 else None
            result = ExamResult.objects.create(
                exam=exam, subject=subject,
                correct_answers=correct, wrong_answers=wrong, blank_answers=blank,
                subject_duration_minutes=duration_val,
            )
            if result.has_errors():
                has_errors = True

        if has_errors:
            messages.success(request, 'Sınav kaydedildi! Hata detaylarını girmek ister misin?')
            return redirect('student:exam_step2', exam_id=exam.id)

        messages.success(request, 'Sınav başarıyla kaydedildi.')
        return redirect('student:dashboard')

    return render(request, 'student/exam_create_step1.html', {'subjects': subjects, 'publishers': publishers})


# ──────────────────────────────────────────────
# STUDENT — EXAM STEP 2 (topic errors)
# ──────────────────────────────────────────────

@student_required
def exam_create_step2(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, student=request.user)

    # All topics as JSON for Alpine.js cascading dropdowns
    topics_qs = Topic.objects.select_related('subject').order_by(
        'subject__name', 'sub_category', 'name'
    )
    topics_json = json.dumps([
        {
            'id': t.id,
            'name': t.name,
            'subject_id': t.subject_id,
            'subject_name': t.subject.name,
            'sub_category': t.sub_category,
        }
        for t in topics_qs
    ])

    subjects = list(Subject.objects.filter(exam_type='TYT').exclude(name='TYT Fen Bilimleri').order_by('name').values('id', 'name'))

    # Step-1 sonuçları: ders bazında Y/B sayısı → bilgi amaçlı gösterim
    step1_results = {
        str(r.subject_id): {'wrong': r.wrong_answers, 'blank': r.blank_answers}
        for r in exam.results.all()
    }

    # Existing errors (for edit mode)
    existing = list(
        exam.topic_errors.select_related('topic__subject')
        .values('topic_id', 'topic__name', 'topic__subject_id',
                'topic__subject__name', 'topic__sub_category',
                'wrong_count', 'blank_count')
    )

    if request.method == 'POST':
        errors_json = request.POST.get('errors_json', '[]')
        try:
            rows = json.loads(errors_json)
        except json.JSONDecodeError:
            rows = []

        ExamTopicError.objects.filter(exam=exam).delete()
        for row in rows:
            topic_id = row.get('topicId')
            wrong = max(0, int(row.get('wrongCount', 0) or 0))
            blank = max(0, int(row.get('blankCount', 0) or 0))
            if topic_id and (wrong > 0 or blank > 0):
                ExamTopicError.objects.update_or_create(
                    exam=exam, topic_id=topic_id,
                    defaults={'wrong_count': wrong, 'blank_count': blank},
                )

        messages.success(request, 'Hata analizi kaydedildi!')
        return redirect('student:exam_detail', exam_id=exam.id)

    return render(request, 'student/exam_create_step2.html', {
        'exam': exam,
        'topics_json': topics_json,
        'subjects_json': json.dumps(subjects),
        'existing_json': json.dumps(existing),
        'step1_results_json': json.dumps(step1_results),
    })


# ──────────────────────────────────────────────
# STUDENT — COMPARE
# ──────────────────────────────────────────────

@student_required
def exam_compare(request):
    raw = request.GET.get('ids', '')
    try:
        ids = [int(i) for i in raw.split(',') if i.strip().isdigit()][:4]
    except ValueError:
        ids = []

    exams = (
        Exam.objects.filter(id__in=ids, student=request.user)
        .prefetch_related('results__subject')
        .order_by('exam_date')
    )

    if exams.count() < 2:
        messages.error(request, 'Karşılaştırma için en az 2 deneme seçin.')
        return redirect('student:dashboard')

    # Build per-exam aggregated result map: {exam_id: {display_subject_name: net|None}}
    result_map = {}
    for exam in exams:
        by_name = {r.subject.name: r.net_score for r in exam.results.all()}
        fen_vals = [by_name[s] for s in _FEN_NAMES if s in by_name]
        result_map[exam.id] = {
            'TYT Türkçe':          by_name.get('TYT Türkçe'),
            'TYT Matematik':       by_name.get('TYT Matematik'),
            'TYT Sosyal Bilimler': by_name.get('TYT Sosyal Bilimler'),
            'TYT Fen Bilimleri':   sum(fen_vals) if fen_vals else None,
        }

    # rows = [{subject, nets: [net_or_none, ...], best}]
    display_subjects = [SimpleNamespace(name=n) for n in _TYT_DISPLAY_SUBJECTS]
    rows = []
    for s in display_subjects:
        nets = [result_map.get(exam.id, {}).get(s.name) for exam in exams]
        valid = [n for n in nets if n is not None]
        best = max(valid) if valid else None
        rows.append({'subject': s, 'nets': nets, 'best': best})

    # Column totals (sum of all 4 display subjects per exam)
    totals = []
    for exam in exams:
        exam_map = result_map.get(exam.id, {})
        totals.append(round(sum(v for v in exam_map.values() if v is not None), 2))

    return render(request, 'student/exam_compare.html', {
        'exams': exams,
        'rows': rows,
        'totals': totals,
    })


# ──────────────────────────────────────────────
# COACH VIEWS
# ──────────────────────────────────────────────

@coach_required
def coach_dashboard(request):
    """DAS-413/414/416: Coach Roster — lists active students with status and quick stats."""
    from users_app.models import User, CoachStudent, CoachAuditLog
    from exams_app.utils import get_student_status

    students = list(
        User.objects.filter(role='student', coach=request.user).order_by('full_name')
    )
    today = date.today()

    # Sync CoachStudent records for all linked students (lazy migration from User.coach FK)
    for student in students:
        CoachStudent.objects.get_or_create(
            coach=request.user, student=student, defaults={'active': True}
        )

    CoachAuditLog.objects.create(coach=request.user, student=None, action='viewed_roster')

    # Pre-fetch primary plan per student (one query, grouped in Python)
    from curriculum_app.models import MacroPlan
    all_plans = list(
        MacroPlan.objects.filter(coach=request.user, student__in=students)
        .order_by('student_id', '-created_at')
        .only('id', 'student_id', 'status', 'sinav_tipi',
              'tyt_start_date', 'tyt_end_date', 'ayt_start_date', 'ayt_end_date',
              'target_date')
    )
    primary_plan_map: dict = {}
    for p in all_plans:
        if p.student_id not in primary_plan_map:
            primary_plan_map[p.student_id] = p

    student_data = []
    status_counts = {'iyi': 0, 'dusus': 0, 'pasif': 0}

    for student in students:
        exams = list(
            Exam.objects.filter(student=student)
            .prefetch_related('results')
            .order_by('-exam_date')[:5]
        )

        last_net = None
        avg_30d = None
        last_exam_ago = None

        if exams:
            last_net = round(float(exams[0].total_net()), 2)
            days_ago = (today - exams[0].exam_date).days
            if days_ago == 0:
                last_exam_ago = 'Bugün'
            elif days_ago == 1:
                last_exam_ago = 'Dün'
            elif days_ago < 15:
                last_exam_ago = f'{days_ago} Gün Önce'
            else:
                last_exam_ago = f'{days_ago}+ Gün'

            since_30 = today - timedelta(days=30)
            exams_30 = list(
                Exam.objects.filter(student=student, exam_date__gte=since_30)
                .prefetch_related('results')
            )
            if exams_30:
                avg_30d = round(
                    sum(float(e.total_net()) for e in exams_30) / len(exams_30), 2
                )

        status_key = get_student_status(student)['status']
        status_counts[status_key] += 1

        initials = ''.join(p[0].upper() for p in student.full_name.split()[:2])

        trend = round(last_net - avg_30d, 2) if last_net is not None and avg_30d is not None else None

        primary_plan = primary_plan_map.get(student.id)
        if primary_plan is None:
            plan_status = 'plan_yok'
        elif primary_plan.status == 'APPROVED':
            plan_status = 'onaylandi'
        else:
            plan_status = 'taslak'

        student_data.append({
            'student': student,
            'last_net': last_net,
            'avg_30d': avg_30d,
            'last_exam_ago': last_exam_ago or '—',
            'status': status_key,
            'initials': initials,
            'trend': trend,
            'primary_plan': primary_plan,
            'plan_status': plan_status,
        })

    # ── Inbox alerts for the slide-out drawer ─────────────────────────────────
    from users_app.models import CoachAlert
    from datetime import date as _date
    from django.db import models as _m
    inbox_alerts = list(
        CoachAlert.objects
        .filter(coach=request.user, is_dismissed=False)
        .filter(_m.Q(expires_at__isnull=True) | _m.Q(expires_at__gte=_date.today()))
        .select_related('student')
        .order_by(
            _m.Case(
                _m.When(severity='critical', then=0),
                _m.When(severity='warning',  then=1),
                _m.When(severity='positive', then=2),
                default=3, output_field=_m.IntegerField(),
            ),
            '-created_at',
        )
    )
    inbox_unread = sum(1 for a in inbox_alerts if not a.is_read)
    inbox_students = list({a.student_id: a.student for a in inbox_alerts}.values())

    return render(request, 'coach/roster.html', {
        'student_data':   student_data,
        'status_counts':  status_counts,
        'total':          len(student_data),
        'inbox_alerts':   inbox_alerts,
        'inbox_unread':   inbox_unread,
        'inbox_students': inbox_students,
        'first_run':      request.session.pop('first_run', False),
    })


@coach_required
def coach_student_detail(request, student_id):
    """DAS-417/418/419: Student Snapshot — coach sees student's dashboard with access guard."""
    from django.http import HttpResponseForbidden
    from users_app.models import User, CoachAuditLog
    from exams_app.utils import get_student_status

    if not coach_can_view_student(request.user, student_id):
        return HttpResponseForbidden('Bu öğrenciye erişim yetkiniz bulunmamaktadır.')

    student = get_object_or_404(User, id=student_id, role='student')
    CoachAuditLog.objects.create(coach=request.user, student=student, action='viewed_snapshot')

    exams = list(
        Exam.objects.filter(student=student)
        .prefetch_related('results__subject')
        .order_by('-exam_date')
    )

    stats = _build_v2_stats(exams, 'all')
    subjects = _build_v2_subjects(exams)
    radar_topics = list(_get_radar_topics(student.id))
    stats['radar_count'] = len(radar_topics)

    status_info = get_student_status(student)

    # ── Trial task panel for the coach to manage this student's tasks ────────
    trial_active_qs = (
        StudentTask.objects.filter(student=student, is_completed=False,
                                   task_source=StudentTask.SOURCE_TRIAL)
        .select_related('topic__subject').order_by('created_at')
    )
    trial_completed_qs = (
        StudentTask.objects.filter(student=student, is_completed=True,
                                   task_source=StudentTask.SOURCE_TRIAL)
        .select_related('topic__subject').order_by('-completed_at')[:30]
    )
    trial_radar_json = json.dumps([
        {'id': t['topic__id'], 'name': t['topic__name'],
         'subject': t['topic__subject__name'], 'sub': t['topic__sub_category'] or '',
         'wrong': t['wrong_total'] or 0, 'blank': t['blank_total'] or 0}
        for t in radar_topics
    ])
    trial_active_json = json.dumps([
        {'id': t.id, 'topic_id': t.topic_id, 'name': t.topic.name,
         'subject': t.topic.subject.name, 'sub': t.topic.sub_category or '',
         'byCoach': t.assigned_by_coach,
         'created_at': _fmt_task_date(t.created_at)}
        for t in trial_active_qs
    ])
    trial_completed_json = json.dumps([
        {'id': t.id, 'topic_id': t.topic_id, 'name': t.topic.name,
         'sub': t.topic.sub_category or '',
         'created_at':   _fmt_task_date(t.created_at),
         'completed_at': _fmt_task_date(t.completed_at),
         'cycle_days':   _task_cycle_days(t.created_at, t.completed_at)}
        for t in trial_completed_qs
    ])

    from users_app.models import StudentAchievement
    recent_badges = [
        {
            'key':   a.badge_key,
            'label': a.get_badge_key_display(),
            'icon':  StudentAchievement.BADGE_META.get(a.badge_key, {}).get('icon', '🏅'),
        }
        for a in StudentAchievement.objects.filter(student=student).order_by('-awarded_at')[:4]
    ]

    return render(request, 'coach/snapshot.html', {
        'student': student,
        'status': status_info['status'],
        'stats': stats,
        'subjects': subjects,
        'exams': exams[:5],
        'trial_radar_json':     trial_radar_json,
        'trial_active_json':    trial_active_json,
        'trial_completed_json': trial_completed_json,
        'student_badges':       recent_badges,
    })


# ──────────────────────────────────────────────
# STUDENT — KENDİME ATA / DİREKT TAMAMLA
# ──────────────────────────────────────────────

@student_required
def self_assign_task(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        data = json.loads(request.body)
        topic_id = int(data['topic_id'])
        direct_complete = bool(data.get('direct_complete', False))
        raw_source = data.get('task_source', StudentTask.SOURCE_TRIAL)
        task_source = raw_source if raw_source in (StudentTask.SOURCE_TRIAL, StudentTask.SOURCE_BRANCH) else StudentTask.SOURCE_TRIAL
    except (KeyError, ValueError, json.JSONDecodeError):
        return JsonResponse({'ok': False}, status=400)

    topic = get_object_or_404(Topic, id=topic_id)
    tomorrow = date.today() + timedelta(days=1)
    if direct_complete:
        task, created = StudentTask.objects.get_or_create(
            student=request.user, topic=topic, task_source=task_source,
            defaults={'assigned_by_coach': False, 'is_completed': True,
                      'completed_at': timezone.now(), 'next_review_date': tomorrow},
        )
        if not created and not task.is_completed:
            task.is_completed = True
            task.completed_at = timezone.now()
            task.next_review_date = tomorrow
            task.save(update_fields=['is_completed', 'completed_at', 'next_review_date'])
        from users_app.services.streak_engine import record_activity
        record_activity(request.user)
    else:
        task, _ = StudentTask.objects.get_or_create(
            student=request.user, topic=topic, task_source=task_source,
            defaults={'assigned_by_coach': False, 'is_completed': False},
        )
    return JsonResponse({'ok': True, 'task_id': task.id})


# ──────────────────────────────────────────────
# COACH — TÜM ÖĞRENCİLERİN SON DENEMELERİ
# ──────────────────────────────────────────────

@coach_required
@coach_required
def unlink_student(request, student_id):
    """AJAX: severs the coach-student relationship without deleting any student data."""
    from django.http import HttpResponseForbidden
    from users_app.models import CoachStudent
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    if not coach_can_view_student(request.user, student_id):
        return HttpResponseForbidden()
    from users_app.models import User as UserModel
    student = get_object_or_404(UserModel, id=student_id, role='student')
    # Sever both the legacy FK and the CoachStudent link table
    if student.coach_id == request.user.id:
        student.coach = None
        student.save(update_fields=['coach'])
    CoachStudent.objects.filter(coach=request.user, student=student).update(active=False)
    return JsonResponse({'ok': True})


def coach_exam_overview(request):
    """Student directory — each student card links to their full exam list."""
    from users_app.models import User
    students = list(User.objects.filter(coach=request.user, role='student'))
    student_data = []
    for s in students:
        exams = list(
            Exam.objects.filter(student=s)
            .prefetch_related('results')
            .order_by('-exam_date')
        )
        last_net = round(float(exams[0].total_net()), 2) if exams else None
        student_data.append({
            'student': s,
            'total': len(exams),
            'last_net': last_net,
            'last_date': exams[0].exam_date if exams else None,
        })
    student_data.sort(key=lambda x: x['last_date'] or date.min, reverse=True)
    return render(request, 'coach/exam_overview.html', {
        'student_data': student_data,
        'v2_shell': True,
        'shell_active': 'Denemeler',
    })


@coach_required
def coach_student_exams(request, student_id):
    """Full trial exam list for a specific student — full feature parity with student Denemeler page."""
    from users_app.models import User
    if not coach_can_view_student(request.user, student_id):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Bu öğrenciye erişim yetkiniz yok.')
    student = get_object_or_404(User, id=student_id, role='student')
    filters = {
        'period':   request.GET.get('period', ''),
        'sort':     request.GET.get('sort', 'date'),
        'has_note': bool(request.GET.get('has_note')),
    }
    ctx = _build_denemeler_v2_context(student, filters)
    ctx.update({
        'student': student,
        'filters': filters,
        'sort_choices': [('date', 'Tarihe Göre'), ('net', 'Net Skora Göre'), ('publisher', 'Yayınevine Göre')],
        'trial_topics': _build_trial_topic_errors(
            student, filters['period'],
            coach_view=True, student_id=student_id,
        ),
        'v2_shell': True,
        'shell_active': 'Denemeler',
    })
    return render(request, 'coach/student_exams.html', ctx)


@coach_required
def coach_student_exam_detail(request, student_id, exam_id):
    """Exam detail view for a coach inspecting a student's specific exam."""
    from users_app.models import User
    if not coach_can_view_student(request.user, student_id):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Bu öğrenciye erişim yetkiniz yok.')
    student = get_object_or_404(User, id=student_id, role='student')
    exam = get_object_or_404(Exam, id=exam_id, student=student)
    results_qs = exam.results.select_related('subject').order_by('subject__name')
    topic_errors_qs = exam.topic_errors.select_related('topic__subject').order_by(
        'topic__subject__name', 'topic__sub_category', 'topic__name'
    )
    ctx = _build_exam_detail_v2_context(exam, results_qs, topic_errors_qs)
    ctx.update({
        'student': student,
        'coach_view': True,
        'v2_shell': True,
        'shell_active': 'Denemeler',
    })
    return render(request, 'coach/student_exam_detail.html', ctx)


# ──────────────────────────────────────────────
# COACH — GÖREV ATA
# ──────────────────────────────────────────────

@coach_required
def assign_task(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        data = json.loads(request.body)
        topic_id = int(data['topic_id'])
        student_id = int(data['student_id'])
    except (KeyError, ValueError, json.JSONDecodeError):
        return JsonResponse({'ok': False}, status=400)

    direct_complete = bool(data.get('direct_complete', False))
    from users_app.models import User
    student = get_object_or_404(User, id=student_id, role='student', coach=request.user)
    topic = get_object_or_404(Topic, id=topic_id)
    tomorrow = date.today() + timedelta(days=1)
    if direct_complete:
        task, created = StudentTask.objects.get_or_create(
            student=student, topic=topic, task_source=StudentTask.SOURCE_TRIAL,
            defaults={'assigned_by_coach': True, 'is_completed': True,
                      'completed_at': timezone.now(), 'next_review_date': tomorrow},
        )
        if not created and not task.is_completed:
            task.is_completed = True
            task.completed_at = timezone.now()
            task.next_review_date = tomorrow
            task.save(update_fields=['is_completed', 'completed_at', 'next_review_date'])
    else:
        task, _ = StudentTask.objects.get_or_create(
            student=student, topic=topic, task_source=StudentTask.SOURCE_TRIAL,
            defaults={'assigned_by_coach': True, 'is_completed': False},
        )
    return JsonResponse({'ok': True, 'task_id': task.id})


# ──────────────────────────────────────────────
# STUDENT — GÖREVİ TAMAMLA
# ──────────────────────────────────────────────

@student_required
def complete_task(request, task_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    task = get_object_or_404(StudentTask, id=task_id, student=request.user)
    task.is_completed = True
    task.completed_at = timezone.now()
    task.next_review_date = date.today() + timedelta(days=1)
    task.save(update_fields=['is_completed', 'completed_at', 'next_review_date'])
    from users_app.services.streak_engine import record_activity
    record_activity(request.user)
    return JsonResponse({'ok': True, 'task': {
        'id': task.id, 'name': task.topic.name,
        'subject': task.topic.subject.name, 'sub': task.topic.sub_category or '',
    }})


# ──────────────────────────────────────────────
# COACH — GÖREVİ TAMAMLA
# ──────────────────────────────────────────────

@coach_required
def coach_complete_task(request, task_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    task = get_object_or_404(StudentTask, id=task_id)
    task.is_completed = True
    task.completed_at = timezone.now()
    task.next_review_date = date.today() + timedelta(days=1)
    task.save(update_fields=['is_completed', 'completed_at', 'next_review_date'])
    return JsonResponse({'ok': True, 'task': {
        'id': task.id, 'name': task.topic.name,
        'subject': task.topic.subject.name, 'sub': task.topic.sub_category or '',
        'byCoach': task.assigned_by_coach,
    }})


# ──────────────────────────────────────────────
# UNIFIED TASK ACTION (cancel / undo / sm2_review)
# ──────────────────────────────────────────────

def _task_action_handler(request, task):
    """Shared logic for cancel_assignment, undo_complete, sm2_review."""
    try:
        data = json.loads(request.body)
        action = data['action']
    except (KeyError, json.JSONDecodeError):
        return JsonResponse({'ok': False}, status=400)

    if action == 'cancel_assignment':
        task.delete()
        return JsonResponse({'ok': True})

    if action == 'undo_complete':
        task.is_completed = False
        task.completed_at = None
        task.next_review_date = None
        task.save(update_fields=['is_completed', 'completed_at', 'next_review_date'])
        return JsonResponse({'ok': True})

    if action == 'sm2_review':
        try:
            quality = int(data['quality_score'])
        except (KeyError, ValueError):
            return JsonResponse({'ok': False}, status=400)
        if quality not in (3, 4, 5):
            return JsonResponse({'ok': False}, status=400)
        from .utils import apply_sm2
        apply_sm2(task, quality)
        return JsonResponse({'ok': True})

    return JsonResponse({'ok': False}, status=400)


@student_required
def student_task_action(request, task_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    task = get_object_or_404(StudentTask, id=task_id, student=request.user)
    return _task_action_handler(request, task)


@coach_required
def coach_task_action(request, task_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    task = get_object_or_404(StudentTask, id=task_id)
    return _task_action_handler(request, task)


# ──────────────────────────────────────────────
# COACH — BRANŞ GÖREV ATA (isolated, branch-only)
# ──────────────────────────────────────────────

@coach_required
def coach_brans_task_assign(request, student_id):
    """Assign or directly complete a BRANCH task for a specific student."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    if not coach_can_view_student(request.user, student_id):
        return JsonResponse({'ok': False}, status=403)
    try:
        data = json.loads(request.body)
        topic_id = int(data['topic_id'])
        direct_complete = bool(data.get('direct_complete', False))
    except (KeyError, ValueError, json.JSONDecodeError):
        return JsonResponse({'ok': False}, status=400)

    from users_app.models import User as UserModel
    student = get_object_or_404(UserModel, id=student_id, role='student')
    topic = get_object_or_404(Topic, id=topic_id)
    tomorrow = date.today() + timedelta(days=1)

    if direct_complete:
        task, created = StudentTask.objects.get_or_create(
            student=student, topic=topic, task_source=StudentTask.SOURCE_BRANCH,
            defaults={'assigned_by_coach': True, 'is_completed': True,
                      'completed_at': timezone.now(), 'next_review_date': tomorrow},
        )
        if not created and not task.is_completed:
            task.is_completed = True
            task.completed_at = timezone.now()
            task.next_review_date = tomorrow
            task.save(update_fields=['is_completed', 'completed_at', 'next_review_date'])
    else:
        task, _ = StudentTask.objects.get_or_create(
            student=student, topic=topic, task_source=StudentTask.SOURCE_BRANCH,
            defaults={'assigned_by_coach': True, 'is_completed': False},
        )
    return JsonResponse({'ok': True, 'task_id': task.id})


@coach_required
def coach_brans_task_action(request, student_id, task_id):
    """Complete / cancel / undo a BRANCH task owned by a specific student."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    if not coach_can_view_student(request.user, student_id):
        return JsonResponse({'ok': False}, status=403)
    from users_app.models import User as UserModel
    student = get_object_or_404(UserModel, id=student_id, role='student')
    task = get_object_or_404(StudentTask, id=task_id, student=student,
                             task_source=StudentTask.SOURCE_BRANCH)
    try:
        data = json.loads(request.body)
        action = data.get('action', '')
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'ok': False}, status=400)

    if action == 'complete':
        task.is_completed = True
        task.completed_at = timezone.now()
        task.next_review_date = date.today() + timedelta(days=1)
        task.save(update_fields=['is_completed', 'completed_at', 'next_review_date'])
        return JsonResponse({'ok': True})
    elif action == 'cancel_assignment':
        task.delete()
        return JsonResponse({'ok': True})
    elif action == 'undo_complete':
        task.is_completed = False
        task.completed_at = None
        task.next_review_date = None
        task.save(update_fields=['is_completed', 'completed_at', 'next_review_date'])
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False, 'error': 'unknown action'}, status=400)


# ──────────────────────────────────────────────
# COACH — DENEME KARŞILAŞTIRMA
# ──────────────────────────────────────────────

@coach_required
def compare_exams(request, student_id):
    from users_app.models import User
    student = get_object_or_404(User, id=student_id, role='student', coach=request.user)

    exam1_id = request.GET.get('exam1', '')
    exam2_id = request.GET.get('exam2', '')
    comparison = None

    if exam1_id and exam2_id:
        e1 = get_object_or_404(Exam, id=exam1_id, student=student)
        e2 = get_object_or_404(Exam, id=exam2_id, student=student)
        comparison = _build_comparison(e1, e2)

    return render(request, 'coach/compare.html', {
        'student': student,
        'comparison': comparison,
    })


# ──────────────────────────────────────────────
# STUDENT — DENEME KARŞILAŞTIRMA
# ──────────────────────────────────────────────

@student_required
def student_compare_exams(request):
    """DAS-372: V1 compare — redirects to V2 when denemeler_v2 flag is on."""
    is_v2 = (
        getattr(request.user, 'denemeler_v2', False)
        or request.GET.get('denemeler_v2') == '1'
        or getattr(settings, 'V2_DEFAULT', False)
    )
    if is_v2:
        return redirect(
            '{}?{}'.format(
                '/student/karsilastir/v2/',
                request.GET.urlencode(),
            )
        )

    student = request.user

    exam1_id = request.GET.get('exam1', '')
    exam2_id = request.GET.get('exam2', '')
    comparison = None

    if exam1_id and exam2_id:
        e1 = get_object_or_404(Exam, id=exam1_id, student=student)
        e2 = get_object_or_404(Exam, id=exam2_id, student=student)
        comparison = _build_comparison(e1, e2)

    return render(request, 'student/compare.html', {'comparison': comparison})


# ──────────────────────────────────────────────
# STUDENT — DENEME KARŞILAŞTIRMA V2 (DAS-361)
# ──────────────────────────────────────────────

_CMP_SUBJECT_MAP = [
    ('TYT Türkçe',          'turkce',  '--turkce'),
    ('TYT Matematik',       'mat',     '--mat'),
    ('TYT Sosyal Bilimler', 'sosyal',  '--sosyal'),
    ('TYT Fen Bilimleri',   'fen',     '--fen'),
]


def _build_v2_chart_data(comparison, exam_a, exam_b):
    """Build SSR-injected JSON payload and template row list for V2 compare."""
    ra = {r.subject.name: r for r in exam_a.results.select_related('subject').all()}
    rb = {r.subject.name: r for r in exam_b.results.select_related('subject').all()}

    net_a_total = round(sum(float(r.net_score) for r in ra.values()), 2)
    net_b_total = round(sum(float(r.net_score) for r in rb.values()), 2)

    labels, deltas, subject_rows = [], [], []

    for delta_row in comparison['deltas']:
        subj_name = delta_row['subject'].name
        _, key, color_var = next(
            (t for t in _CMP_SUBJECT_MAP if t[0] == subj_name),
            ('', subj_name, '--primary'),
        )
        display = (
            delta_row['subject'].display_name
            if hasattr(delta_row['subject'], 'display_name')
            else (subj_name[4:] if subj_name[:4] == 'TYT ' else subj_name)
        )
        labels.append(display)
        deltas.append(delta_row['net_delta'])

        # Per-subject D/Y/B — aggregate Fen split subjects
        if subj_name == 'TYT Fen Bilimleri':
            fen_ra = [r for n, r in ra.items() if n in _FEN_NAMES]
            fen_rb = [r for n, r in rb.items() if n in _FEN_NAMES]
            ca = sum(r.correct_answers for r in fen_ra)
            ya = sum(r.wrong_answers   for r in fen_ra)
            ba = sum(r.blank_answers   for r in fen_ra)
            cb = sum(r.correct_answers for r in fen_rb)
            yb = sum(r.wrong_answers   for r in fen_rb)
            bb = sum(r.blank_answers   for r in fen_rb)
        else:
            r_a = ra.get(subj_name)
            r_b = rb.get(subj_name)
            ca = r_a.correct_answers if r_a else 0
            ya = r_a.wrong_answers   if r_a else 0
            ba = r_a.blank_answers   if r_a else 0
            cb = r_b.correct_answers if r_b else 0
            yb = r_b.wrong_answers   if r_b else 0
            bb = r_b.blank_answers   if r_b else 0

        subject_rows.append({
            'name':      display,
            'color_var': color_var,
            'net_a':     delta_row['net_a'],
            'net_b':     delta_row['net_b'],
            'delta':     delta_row['net_delta'],
            'correct_a': ca, 'wrong_a': ya, 'blank_a': ba,
            'correct_b': cb, 'wrong_b': yb, 'blank_b': bb,
        })

    chart_json = json.dumps({
        'labels':      labels,
        'deltas':      deltas,
        'net_a':       net_a_total,
        'net_b':       net_b_total,
        'total_delta': round(net_b_total - net_a_total, 2),
    })
    chart_data = SimpleNamespace(
        net_a=net_a_total, net_b=net_b_total,
        total_delta=round(net_b_total - net_a_total, 2),
    )
    return chart_json, chart_data, subject_rows


@student_required
def student_compare_v2(request):
    """DAS-361: V2 mobile compare page — diverging-bar + desktop radar preservation."""
    is_v2 = (
        getattr(request.user, 'denemeler_v2', False)
        or request.GET.get('denemeler_v2') == '1'
        or getattr(settings, 'V2_DEFAULT', False)
    )
    if not is_v2:
        return redirect(
            '{}?{}'.format(
                '/student/karsilastir/',
                request.GET.urlencode(),
            )
        )

    exam1_id = request.GET.get('exam1', '')
    exam2_id = request.GET.get('exam2', '')
    comparison = chart_json = chart_data = None
    subject_rows = []

    if exam1_id and exam2_id:
        e1 = get_object_or_404(Exam, id=exam1_id, student=request.user)
        e2 = get_object_or_404(Exam, id=exam2_id, student=request.user)
        comparison = _build_comparison(e1, e2)
        chart_json, chart_data, subject_rows = _build_v2_chart_data(
            comparison, comparison['exam_a'], comparison['exam_b']
        )

    return render(request, 'student/karsilastir_v2.html', {
        'comparison':   comparison,
        'chart_json':   chart_json,
        'chart_data':   chart_data,
        'subject_rows': subject_rows,
        'v2_shell':     True,
    })


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────

def _build_comparison(exam_1, exam_2):
    """Build radar/delta/closed-gaps data. Older exam → A (baseline), newer → B (current)."""
    exam_a, exam_b = (exam_1, exam_2) if exam_1.exam_date <= exam_2.exam_date else (exam_2, exam_1)

    results_a = list(ExamResult.objects.filter(exam=exam_a).select_related('subject'))
    results_b = list(ExamResult.objects.filter(exam=exam_b).select_related('subject'))

    def _by_name(results):
        return {r.subject.name: r for r in results}

    def _fen_net(by_name):
        return sum(float(r.net_score) for name, r in by_name.items() if name in _FEN_NAMES)

    def _fen_dur(by_name):
        durs = [r.subject_duration_minutes for name, r in by_name.items()
                if name in _FEN_NAMES and r.subject_duration_minutes is not None]
        return sum(durs) if durs else None

    map_a, map_b = _by_name(results_a), _by_name(results_b)

    deltas = []
    for name in _TYT_DISPLAY_SUBJECTS:
        if name == 'TYT Fen Bilimleri':
            net_a  = _fen_net(map_a)
            net_b  = _fen_net(map_b)
            dur_a  = _fen_dur(map_a)
            dur_b  = _fen_dur(map_b)
            subj   = SimpleNamespace(name='TYT Fen Bilimleri')
        else:
            r_a = map_a.get(name)
            r_b = map_b.get(name)
            net_a = float(r_a.net_score) if r_a else 0.0
            net_b = float(r_b.net_score) if r_b else 0.0
            dur_a = r_a.subject_duration_minutes if r_a else None
            dur_b = r_b.subject_duration_minutes if r_b else None
            real  = r_a or r_b
            subj  = real.subject if real else SimpleNamespace(name=name)
        deltas.append({
            'subject': subj,
            'net_a': net_a, 'net_b': net_b,
            'net_delta': round(net_b - net_a, 2),
            'dur_a': dur_a, 'dur_b': dur_b,
            'dur_delta': (dur_b - dur_a) if (dur_a is not None and dur_b is not None) else None,
        })

    errors_a = set(ExamTopicError.objects.filter(exam=exam_a).values_list('topic_id', flat=True))
    errors_b = set(ExamTopicError.objects.filter(exam=exam_b).values_list('topic_id', flat=True))
    closed_gap_ids = errors_a - errors_b
    closed_gaps = (
        Topic.objects.filter(id__in=closed_gap_ids).select_related('subject')
        .order_by('subject__name', 'name')
    ) if closed_gap_ids else []

    def _s(n):
        return n[4:] if n[:4] in ('TYT ', 'AYT ') else n

    return {
        'exam_a': exam_a,
        'exam_b': exam_b,
        'deltas': deltas,
        'closed_gaps': closed_gaps,
        'radar_json': json.dumps({
            'labels': [_s(d['subject'].name) for d in deltas],
            'data_a': [d['net_a'] for d in deltas],
            'data_b': [d['net_b'] for d in deltas],
            'label_a': exam_a.custom_name,
            'label_b': exam_b.custom_name,
        }),
    }


def _get_radar_topics(student_id, since=None):
    """Gray topics that don't yet have a TRIAL StudentTask for this student."""
    all_gray = list(_get_gray_topics(student_id, since=since))
    assigned_ids = set(
        StudentTask.objects.filter(student_id=student_id,
                                   task_source=StudentTask.SOURCE_TRIAL)
        .values_list('topic_id', flat=True)
    )
    return [t for t in all_gray if t['topic__id'] not in assigned_ids]


def _get_gray_topics(student_id, last_n=None, since=None):
    from django.db.models import Count, Sum
    qs = ExamTopicError.objects.filter(exam__student_id=student_id)

    if since:
        qs = qs.filter(exam__exam_date__gte=since)
    elif last_n:
        recent_ids = (
            Exam.objects.filter(student_id=student_id)
            .order_by('-exam_date')
            .values_list('id', flat=True)[:last_n]
        )
        qs = qs.filter(exam_id__in=recent_ids)

    return (
        qs.values('topic__id', 'topic__name', 'topic__subject__name', 'topic__sub_category')
        .annotate(
            error_count=Sum('wrong_count') + Sum('blank_count'),
            wrong_total=Sum('wrong_count'),
            blank_total=Sum('blank_count'),
        )
        .order_by('-error_count')[:20]
    )


def _build_student_chart_data(student_id, since=None):
    qs = Exam.objects.filter(student_id=student_id)
    if since:
        qs = qs.filter(exam_date__gte=since)
    exams = list(qs.prefetch_related('results__subject').order_by('exam_date'))

    if not exams:
        return {}

    labels, exam_ids, exam_names, totals = [], [], [], []
    subject_nets = {name: [] for name in _TYT_DISPLAY_SUBJECTS}

    for exam in exams:
        labels.append(str(exam.exam_date))
        by_name = {name: 0 for name in _TYT_DISPLAY_SUBJECTS}
        exam_ids.append(exam.id)
        exam_names.append(exam.custom_name)
        result_map = {r.subject.name: r.net_score for r in exam.results.all()}
        result_map['TYT Fen Bilimleri'] = sum(result_map.get(s, 0) for s in _FEN_NAMES)
        for name in _TYT_DISPLAY_SUBJECTS:
            by_name[name] = result_map.get(name, 0)

        total = 0
        for name in _TYT_DISPLAY_SUBJECTS:
            net = by_name.get(name, 0)
            subject_nets[name].append(net)
            total += net
        totals.append(round(total, 2))

    subjects_data = []
    for i, name in enumerate(_TYT_DISPLAY_SUBJECTS):
        colors = SUBJECT_COLORS.get(name, DEFAULT_COLORS[i % len(DEFAULT_COLORS)])
        subjects_data.append({'name': name, 'border': colors['border'], 'bg': colors['background'], 'data': subject_nets[name]})

    return {'labels': labels, 'exam_ids': exam_ids, 'exam_names': exam_names, 'subjects': subjects_data, 'totals': totals}


def _build_subject_duration_scatter_data(student_id, since=None):
    from collections import defaultdict
    qs = (
        ExamResult.objects
        .filter(exam__student_id=student_id, subject_duration_minutes__isnull=False)
        .select_related('exam', 'exam__publisher', 'subject')
    )
    if since:
        qs = qs.filter(exam__exam_date__gte=since)

    # Collect raw results grouped by subject name, keeping exam_id for Fen aggregation
    raw = defaultdict(list)
    for r in qs:
        raw[r.subject.name].append({
            'exam_id':   r.exam.id,
            'x':         r.subject_duration_minutes,
            'y':         r.net_score,
            'exam':      r.exam.custom_name,
            'publisher': r.exam.publisher.name,
        })

    if not raw:
        return {'datasets': []}

    # Aggregate all Fen variants into one point per exam (sum duration + sum net)
    fen_by_exam = defaultdict(lambda: {'x': 0, 'y': 0.0, 'exam': None, 'publisher': None})
    for name in _FEN_NAMES:
        for pt in raw.get(name, []):
            e = fen_by_exam[pt['exam_id']]
            e['x'] += pt['x']
            e['y'] += pt['y']
            e['exam'] = pt['exam']
            e['publisher'] = pt['publisher']

    # Build final grouped dict aligned to _TYT_DISPLAY_SUBJECTS
    grouped = {}
    for name in _TYT_DISPLAY_SUBJECTS:
        if name == 'TYT Fen Bilimleri':
            pts = [{'x': v['x'], 'y': v['y'], 'exam': v['exam'], 'publisher': v['publisher']}
                   for v in fen_by_exam.values() if v['x'] > 0]
            if pts:
                grouped[name] = pts
        else:
            pts = [{'x': p['x'], 'y': p['y'], 'exam': p['exam'], 'publisher': p['publisher']}
                   for p in raw.get(name, [])]
            if pts:
                grouped[name] = pts

    if not grouped:
        return {'datasets': []}

    datasets = []
    for i, name in enumerate(_TYT_DISPLAY_SUBJECTS):
        if name not in grouped:
            continue
        colors = SUBJECT_COLORS.get(name, DEFAULT_COLORS[i % len(DEFAULT_COLORS)])
        datasets.append({
            'label':  name[4:],
            'border': colors['border'],
            'data':   grouped[name],
        })

    return {'datasets': datasets}


# ── Branş Deneme views ─────────────────────────────────────────────────────────

def _build_brans_topic_analytics(student, since=None, top_n=10, ders_id=None):
    """
    Return top-N most-wrong topics for a student across all BransDeneme entries.
    Mirrors _get_gray_topics() but aggregates BransTopicError.yanlis_sayisi.
    Includes last-seen date and optional ders filter.
    """
    from django.db.models import Sum, Max
    qs = BransTopicError.objects.filter(brans_deneme__student=student)
    if since:
        qs = qs.filter(brans_deneme__tarih__gte=since)
    if ders_id:
        qs = qs.filter(brans_deneme__ders_id=ders_id)
    return list(
        qs.values(
            'topic__id', 'topic__name',
            'topic__subject__name', 'topic__subject__exam_type',
            'topic__subject__id',
            'topic__sub_category',
        )
        .annotate(
            toplam_yanlis=Sum('yanlis_sayisi'),
            son_tarih=Max('brans_deneme__tarih'),
        )
        .order_by('-toplam_yanlis')[:top_n]
    )


def _build_brans_chart_data(entries):
    """Build Chart.js dataset for BransDeneme net trend, one line per ders."""
    from collections import defaultdict
    by_ders = defaultdict(list)
    for e in entries:
        by_ders[e.ders].append(e)

    # Collect all unique dates sorted ascending
    all_dates = sorted({e.tarih.isoformat() for e in entries})
    if not all_dates:
        return {'labels': [], 'datasets': []}

    color_cycle = [
        {'border': '#3b82f6', 'bg': 'rgba(59,130,246,0.12)'},
        {'border': '#8b5cf6', 'bg': 'rgba(139,92,246,0.12)'},
        {'border': '#10b981', 'bg': 'rgba(16,185,129,0.12)'},
        {'border': '#f59e0b', 'bg': 'rgba(245,158,11,0.12)'},
        {'border': '#ec4899', 'bg': 'rgba(236,72,153,0.12)'},
        {'border': '#6366f1', 'bg': 'rgba(99,102,241,0.12)'},
    ]
    datasets = []
    for i, (ders, ders_entries) in enumerate(by_ders.items()):
        net_by_date = {}
        for e in ders_entries:
            # Keep the best net for that date if multiple entries exist
            d = e.tarih.isoformat()
            if d not in net_by_date or e.net > net_by_date[d]:
                net_by_date[d] = e.net
        color = color_cycle[i % len(color_cycle)]
        datasets.append({
            'label': ders.display_name,
            'data': [net_by_date.get(d) for d in all_dates],
            'borderColor': color['border'],
            'backgroundColor': color['bg'],
            'tension': 0.3,
            'spanGaps': True,
        })

    return {'labels': all_dates, 'datasets': datasets}


def _build_brans_compare_context(entry_a, entry_b):
    """Shared comparison context for student and coach brans compare views."""
    if entry_a.tarih > entry_b.tarih:
        entry_a, entry_b = entry_b, entry_a

    chart_data = json.dumps({
        'labels': [entry_a.tarih.isoformat(), entry_b.tarih.isoformat()],
        'datasets': [{
            'label': 'Net',
            'data':  [entry_a.net, entry_b.net],
            'backgroundColor': ['rgba(99,102,241,0.6)', 'rgba(16,185,129,0.6)'],
            'borderColor':     ['#6366f1', '#10b981'],
            'borderWidth': 1,
        }],
    })

    te_a = list(entry_a.topic_errors.select_related('topic'))
    te_b = list(entry_b.topic_errors.select_related('topic'))
    errors_a = {te.topic_id: te.yanlis_sayisi for te in te_a}
    errors_b = {te.topic_id: te.yanlis_sayisi for te in te_b}
    topic_name = {te.topic_id: te.topic.name for te in te_a + te_b}
    topic_deltas = sorted(
        [
            {
                'name':  name,
                'a_val': errors_a.get(tid, 0),
                'b_val': errors_b.get(tid, 0),
                'delta': errors_b.get(tid, 0) - errors_a.get(tid, 0),
            }
            for tid, name in topic_name.items()
        ],
        key=lambda x: abs(x['delta']),
        reverse=True,
    )

    return {
        'a': entry_a, 'b': entry_b,
        'a_b_pairs': [
            (entry_a, 'Eski', 'bg-gray-50'),
            (entry_b, 'Yeni', 'bg-indigo-50'),
        ],
        'deltas': [
            ('Net',    round(entry_b.net    - entry_a.net,    2), True),
            ('Doğru',  entry_b.dogru  - entry_a.dogru,           True),
            ('Yanlış', entry_b.yanlis - entry_a.yanlis,          False),
            ('Boş',    entry_b.bos    - entry_a.bos,             False),
        ],
        'chart_data_json': chart_data,
        'topic_deltas':    topic_deltas,
    }


@student_required
def brans_create(request):
    """Create a new branch exam entry for the current student."""
    has_next = bool(request.POST.get('next') or request.GET.get('next'))
    next_url = _safe_next_url(request, '/brans/') if has_next else ''
    if request.method == 'POST':
        form = BransDenemeForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.student = request.user
            entry.save()
            from users_app.services.streak_engine import record_activity
            record_activity(request.user)
            if has_next:
                topic_url = reverse('student:brans_topic_errors', kwargs={'pk': entry.pk})
                return redirect(_append_next(topic_url, next_url))
            messages.success(request, 'Branş denemesi kaydedildi. Konu hatalarını girebilirsiniz.')
            return redirect('student:brans_topic_errors', pk=entry.pk)
    else:
        form = BransDenemeForm()

    subjects_json = json.dumps({str(s.pk): s.question_count for s in Subject.objects.only('pk', 'question_count')})
    return render(request, 'student/brans_create.html', {
        'form': form,
        'subjects_json': subjects_json,
        'next_url': next_url,
        'back_url': next_url,
        'page_title': 'Branş Denemesi Ekle',
        'submit_label': 'Branş Denemesini Kaydet',
    })


@student_required
def brans_detail(request, pk):
    entry = get_object_or_404(
        BransDeneme.objects.select_related('ders'), pk=pk, student=request.user
    )
    topic_errors = entry.topic_errors.select_related('topic').order_by('-yanlis_sayisi')

    prev = (
        BransDeneme.objects
        .filter(student=request.user, ders=entry.ders, tarih__lt=entry.tarih)
        .order_by('-tarih', '-olusturulma_tarihi')
        .first()
    )
    delta = None
    if prev:
        delta = {
            'net':        round(entry.net - prev.net, 2),
            'dogru':      entry.dogru - prev.dogru,
            'yanlis':     entry.yanlis - prev.yanlis,
            'bos':        entry.bos - prev.bos,
            'prev_tarih': prev.tarih,
            'prev_net':   prev.net,
        }

    return render(request, 'student/brans_detail.html', {
        'entry': entry,
        'topic_errors': topic_errors,
        'delta': delta,
    })


@student_required
def brans_edit(request, pk):
    entry = get_object_or_404(BransDeneme, pk=pk, student=request.user)
    if request.method == 'POST':
        form = BransDenemeForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            messages.success(request, 'Branş denemesi güncellendi. Konu hatalarını düzenleyebilirsiniz.')
            return redirect('student:brans_topic_errors', pk=entry.pk)
    else:
        form = BransDenemeForm(instance=entry)
    subjects_json = json.dumps({str(s.pk): s.question_count for s in Subject.objects.only('pk', 'question_count')})
    return render(request, 'student/brans_edit.html', {'form': form, 'entry': entry, 'subjects_json': subjects_json})


@student_required
def brans_delete(request, pk):
    entry = get_object_or_404(BransDeneme, pk=pk, student=request.user)
    if request.method == 'POST':
        entry.delete()
        messages.success(request, 'Branş denemesi silindi.')
    return redirect('brans_hub')


@coach_required
def coach_brans_analytics(request):
    """Phase 7: student router — cards linking to each student's branch analytics hub."""
    from django.db.models import Avg, Count, Max
    from users_app.models import User as UserModel

    students = list(
        UserModel.objects
        .filter(role='student', coach=request.user)
        .order_by('full_name')
    )

    student_data = []
    for s in students:
        stats = (
            BransDeneme.objects
            .filter(student=s)
            .aggregate(total=Count('id'), last_date=Max('tarih'), avg_net=Avg('net'))
        )
        student_data.append({
            'student':  s,
            'total':    stats['total'] or 0,
            'last_date': stats['last_date'],
            'avg_net':  round(float(stats['avg_net']), 1) if stats['avg_net'] else None,
        })

    return render(request, 'coach/brans_analytics.html', {
        'student_data': student_data,
        'v2_shell':     True,
        'shell_active': 'Branş',
    })


@coach_required
def coach_brans_student_detail(request, student_id):
    """Drill-down: coach sees one student's BransDeneme entries with chart + period filter."""
    from users_app.models import User as UserModel
    from collections import defaultdict

    student = get_object_or_404(UserModel, id=student_id, role='student', coach=request.user)

    period = request.GET.get('period', 'all')
    since = None
    if period == '30':
        since = date.today() - timedelta(days=30)
    elif period == '90':
        since = date.today() - timedelta(days=90)

    ders_filter = request.GET.get('ders', '')
    ders_filter_id = None

    qs = BransDeneme.objects.filter(student=student).select_related('ders')
    if since:
        qs = qs.filter(tarih__gte=since)
    if ders_filter:
        try:
            ders_filter_id = int(ders_filter)
            qs = qs.filter(ders_id=ders_filter_id)
        except ValueError:
            ders_filter = ''

    entries = list(qs.order_by('tarih', 'ders__exam_type', 'ders__name'))

    grouped = defaultdict(list)
    for entry in entries:
        grouped[entry.ders].append(entry)
    for k in grouped:
        grouped[k].sort(key=lambda e: e.tarih, reverse=True)

    chart_data = _build_brans_chart_data(entries)
    top_topics = _build_brans_topic_analytics(student, since=since, ders_id=ders_filter_id)
    all_ders = (
        Subject.objects
        .filter(brans_denemeler__student=student)
        .distinct()
        .order_by('exam_type', 'name')
    )

    return render(request, 'coach/brans_student_detail.html', {
        'student': student,
        'grouped': dict(grouped),
        'chart_data_json': json.dumps(chart_data),
        'has_entries': bool(entries),
        'period': period,
        'period_choices': [('all', 'Tümü'), ('90', '3 Ay'), ('30', '30 Gün')],
        'top_topics': top_topics,
        'all_ders': all_ders,
        'ders_filter': ders_filter,
    })


@coach_required
def coach_brans_compare(request, student_id):
    """Coach compares two BransDeneme entries for one of their students."""
    from users_app.models import User as UserModel

    student = get_object_or_404(UserModel, id=student_id, role='student', coach=request.user)

    try:
        a_id = int(request.GET['a'])
        b_id = int(request.GET['b'])
    except (KeyError, ValueError):
        return redirect('coach:brans_student_detail', student_id=student_id)

    entry_a = get_object_or_404(BransDeneme.objects.select_related('ders'), pk=a_id, student=student)
    entry_b = get_object_or_404(BransDeneme.objects.select_related('ders'), pk=b_id, student=student)

    ctx = _build_brans_compare_context(entry_a, entry_b)
    ctx['student'] = student
    return render(request, 'coach/brans_compare.html', ctx)


def _render_brans_topic_errors(request, entry, *, final_fallback, back_fallback):
    """Enter / edit per-topic wrong counts for a BransDeneme entry."""
    next_url = _safe_next_url(request, final_fallback)
    has_next = bool(request.POST.get('next') or request.GET.get('next'))
    back_url = next_url if has_next else back_fallback

    # All topics for the entry's subject, as JSON for the Alpine UI
    topics_qs = Topic.objects.filter(subject=entry.ders).order_by('sub_category', 'name')
    topics_json = json.dumps([
        {
            'id': t.id,
            'name': t.name,
            'sub_category': t.sub_category,
        }
        for t in topics_qs
    ])

    # Existing errors pre-populate the form on re-edit
    existing_json = json.dumps([
        {'topicId': e.topic_id, 'topicName': e.topic.name,
         'subCat': e.topic.sub_category, 'yanlisSayisi': e.yanlis_sayisi}
        for e in entry.topic_errors.select_related('topic').order_by('topic__sub_category', 'topic__name')
    ])

    if request.method == 'POST':
        raw = request.POST.get('errors_json', '[]')
        try:
            rows = json.loads(raw)
        except json.JSONDecodeError:
            rows = []

        valid_topic_ids = set(
            Topic.objects.filter(subject=entry.ders).values_list('id', flat=True)
        )
        total_yanlis = sum(max(0, int(r.get('yanlisSayisi', 0) or 0)) for r in rows)
        if total_yanlis > entry.yanlis:
            messages.error(
                request,
                f'Takip edilen toplam yanlış ({total_yanlis}) '
                f'deneme yanlış sayısını ({entry.yanlis}) aşamaz.'
            )
        else:
            BransTopicError.objects.filter(brans_deneme=entry).delete()
            for row in rows:
                topic_id = row.get('topicId')
                yanlis   = max(0, int(row.get('yanlisSayisi', 0) or 0))
                if topic_id and yanlis > 0 and topic_id in valid_topic_ids:
                    BransTopicError.objects.update_or_create(
                        brans_deneme=entry, topic_id=topic_id,
                        defaults={'yanlis_sayisi': yanlis},
                    )
            messages.success(request, 'Konu hataları kaydedildi.')
            return redirect(next_url if has_next else final_fallback)

    return render(request, 'student/brans_topic_errors.html', {
        'entry': entry,
        'topics_json': topics_json,
        'existing_json': existing_json,
        'next_url': next_url if has_next else '',
        'back_url': back_url,
    })


@student_required
def dashboard_chart_data(request):
    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return redirect('student:dashboard')

    chart_period, chart_since, chart_period_label = _chart_period_since(request.GET.get('chart_period', '30'))
    chart_data = _build_student_chart_data(request.user.id, since=chart_since)
    scatter_data = _build_subject_duration_scatter_data(request.user.id, since=chart_since)
    return JsonResponse({
        'chart_data': chart_data,
        'scatter_data': scatter_data,
        'has_exams': bool(chart_data.get('labels')),
        'chart_period': chart_period,
        'chart_period_label': chart_period_label,
    })


@student_required
def brans_topic_errors(request, pk):
    entry = get_object_or_404(BransDeneme.objects.select_related('ders'), pk=pk, student=request.user)
    return _render_brans_topic_errors(
        request,
        entry,
        final_fallback=reverse('student:brans_detail', kwargs={'pk': entry.pk}),
        back_fallback=reverse('student:brans_detail', kwargs={'pk': entry.pk}),
    )


@coach_required
def coach_brans_topic_errors(request, student_id, pk):
    if not coach_can_view_student(request.user, student_id):
        return HttpResponseForbidden('Bu öğrenciyi görüntüleme yetkiniz yok.')
    entry = get_object_or_404(
        BransDeneme.objects.select_related('ders', 'student'),
        pk=pk,
        student_id=student_id,
    )
    return _render_brans_topic_errors(
        request,
        entry,
        final_fallback=reverse('coach:student_brans_hub', kwargs={'student_id': student_id}),
        back_fallback=reverse('coach:brans_student_detail', kwargs={'student_id': student_id}),
    )
