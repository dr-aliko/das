import io
import json
from datetime import date

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views import View

from users_app.decorators import coach_required, student_required
from tasks_app.forms import GorevForm
from tasks_app.models import AktiviteTipi
from tasks_app.services import api_client, tasks, week as week_svc
from tasks_app.services import resources, students


def _body(request) -> dict:
    try:
        return json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        return {}


def _build_meta(aktivite_tipi: str, body: dict) -> dict | None:
    """
    meta is only set for konu_anlatimi.
    meta.videos is the source of truth for edit hydration — never re-fetched from API.
    When youtube_playlist_pk is present the source is a local YouTubePlaylist row.
    """
    if aktivite_tipi != AktiviteTipi.KONU_ANLATIMI:
        return None
    yt_pk = body.get("youtube_playlist_pk")
    if yt_pk:
        return {
            "source":              "youtube",
            "youtube_playlist_pk": int(yt_pk),
            "sinav_tipi":          body.get("sinav_tipi"),
            "liste_baslik":        body.get("liste_baslik", ""),
            "videos":              body.get("videos", []),
        }
    return {
        "sinav_tipi":   body.get("sinav_tipi"),
        "ders_id":      body.get("konu_ders_id"),
        "liste_id":     body.get("liste_id"),
        "liste_baslik": body.get("liste_baslik", ""),  # playlist name for student detail view
        "videos":       body.get("videos", []),  # [{id, title, duration}]
    }


# ── Main page ─────────────────────────────────────────────────────────────────

@method_decorator(coach_required, name='dispatch')
class HaftaView(View):
    def get(self, request):
        coached_students = students.list_for_coach(request.user)
        return render(request, "coach/tasks/hafta.html", {"students": coached_students})


# ── Students (read-only) ──────────────────────────────────────────────────────

@method_decorator(coach_required, name='dispatch')
class StudentListView(View):
    def get(self, request):
        data = [
            {"id": s.id, "full_name": s.full_name}
            for s in students.list_for_coach(request.user)
        ]
        return JsonResponse({"students": data})


# ── Görevler ─────────────────────────────────────────────────────────────────

@method_decorator(coach_required, name='dispatch')
class GorevListView(View):
    def get(self, request):
        try:
            student_id = int(request.GET["student_id"])
            hafta = date.fromisoformat(request.GET["hafta"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "student_id ve hafta=YYYY-MM-DD gerekli"}, status=400)
        basi, sonu = week_svc.week_bounds(hafta)
        gorevler = tasks.week_for_student(request.user, student_id, basi, sonu)
        return JsonResponse({
            "hafta_basi": basi.isoformat(),
            "hafta_sonu": sonu.isoformat(),
            "gorevler": gorevler,
            "gunluk_toplamlar": week_svc.daily_totals(gorevler),
        })

    def post(self, request):
        body = _body(request)
        form = GorevForm(body)
        if not form.is_valid():
            return JsonResponse({"errors": form.errors}, status=400)
        cd = form.cleaned_data
        student_id = cd.get("student_id") or body.get("student_id")
        if not student_id:
            return JsonResponse({"errors": "student_id gerekli"}, status=400)
        grup = tasks.create(
            request.user,
            int(student_id),
            tarih=cd["tarih"],
            ders_title=cd["ders_title"],
            ozel_sure_dk=cd.get("ozel_sure_dk"),
            aktivite_tipi=cd["aktivite_tipi"],
            detaylar=cd["detaylar"],
            meta=_build_meta(cd["aktivite_tipi"], body),
        )
        if not grup:
            return JsonResponse({"errors": "Öğrenci bulunamadı"}, status=404)
        return JsonResponse({"id": grup.id}, status=201)


@method_decorator(coach_required, name='dispatch')
class GorevDetailView(View):
    def get(self, request, pk: int):
        data = tasks.get_one(request.user, pk)
        if not data:
            return JsonResponse({"errors": "Bulunamadı"}, status=404)
        return JsonResponse(data)

    def put(self, request, pk: int):
        body = _body(request)
        form = GorevForm(body)
        if not form.is_valid():
            return JsonResponse({"errors": form.errors}, status=400)
        cd = form.cleaned_data
        ok = tasks.update(
            request.user,
            pk,
            tarih=cd["tarih"],
            ders_title=cd["ders_title"],
            ozel_sure_dk=cd.get("ozel_sure_dk"),
            aktivite_tipi=cd["aktivite_tipi"],
            detaylar=cd["detaylar"],
            meta=_build_meta(cd["aktivite_tipi"], body),
        )
        return JsonResponse({"ok": ok}, status=200 if ok else 404)

    def delete(self, request, pk: int):
        ok = tasks.delete(request.user, pk)
        return JsonResponse({"ok": ok}, status=200 if ok else 404)


@method_decorator(coach_required, name='dispatch')
class GorevCopyView(View):
    def post(self, request, pk: int):
        try:
            hedef = date.fromisoformat(_body(request)["hedef_tarih"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "hedef_tarih gerekli"}, status=400)
        yeni = tasks.copy_to_date(request.user, pk, hedef)
        if not yeni:
            return JsonResponse({"ok": False}, status=404)
        return JsonResponse({"id": yeni.id}, status=201)


@method_decorator(coach_required, name='dispatch')
class GorevReorderView(View):
    def post(self, request, pk: int):
        try:
            hedef_id = int(_body(request)["hedef_id"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "hedef_id gerekli"}, status=400)
        ok = tasks.swap_order(request.user, pk, hedef_id)
        return JsonResponse({"ok": ok}, status=200 if ok else 400)


# ── Kaynak kitaplar ───────────────────────────────────────────────────────────

@method_decorator(coach_required, name='dispatch')
class KaynakKitapListView(View):
    def get(self, request):
        try:
            student_id = int(request.GET["student_id"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "student_id gerekli"}, status=400)
        ders = request.GET.get("ders", "")
        sinav_tipi = request.GET.get("sinav_tipi", "")
        kitaplar = resources.list_for_student(request.user, student_id, ders, sinav_tipi)
        return JsonResponse({"kitaplar": kitaplar})

    def post(self, request):
        body = _body(request)
        ok = resources.add(
            request.user,
            body.get("student_id"),
            body.get("kitap_adi", ""),
            body.get("ders", ""),
            body.get("sinav_tipi", ""),
        )
        return JsonResponse({"ok": ok}, status=201 if ok else 400)


# ── External API proxies ──────────────────────────────────────────────────────

@method_decorator(coach_required, name='dispatch')
class DerslerView(View):
    def get(self, request):
        return JsonResponse({"dersler": api_client.get_dersler()})


@method_decorator(coach_required, name='dispatch')
class OynatmaListeleriView(View):
    def get(self, request, ders_id: int):
        tip = request.GET.get("tip", "")
        return JsonResponse({"listeler": api_client.get_oynatma_listeleri(ders_id, tip)})


@method_decorator(coach_required, name='dispatch')
class VideolarView(View):
    def get(self, request, liste_id: int):
        return JsonResponse({"videolar": api_client.get_videolar(liste_id)})


# ── YouTube playlist endpoints ────────────────────────────────────────────────

@method_decorator(coach_required, name='dispatch')
class YoutubePlaylistPreviewView(View):
    def post(self, request):
        from tasks_app.models import YouTubePlaylist
        from tasks_app.services.youtube_playlist_importer import (
            extract_playlist_id, fetch_playlist_meta, detect_subject_and_exam_type,
            InvalidPlaylistURL, PlaylistNotFound, YouTubeAPIError,
        )
        body = _body(request)
        url = (body.get("url") or "").strip()
        pid = extract_playlist_id(url)
        if not pid:
            return JsonResponse({"error": "Geçerli bir YouTube oynatma listesi URL'si girin."}, status=400)

        try:
            meta = fetch_playlist_meta(pid)
        except PlaylistNotFound:
            return JsonResponse({"error": "Bu playlist bulunamadı veya gizli."}, status=404)
        except YouTubeAPIError as exc:
            return JsonResponse({"error": str(exc)}, status=502)

        subject, exam_type = detect_subject_and_exam_type(meta['title'])
        existing = YouTubePlaylist.objects.filter(playlist_id=pid).first()

        from exams_app.models import Subject
        subj_qs = Subject.objects.order_by('exam_type', 'name')
        if exam_type:
            subj_qs = subj_qs.filter(exam_type=exam_type)
        subjects = [{'id': s.pk, 'display_name': s.display_name, 'exam_type': s.exam_type} for s in subj_qs]

        return JsonResponse({
            "playlist_id":           pid,
            "title":                 meta['title'],
            "channel_title":         meta['channel_title'],
            "video_count":           meta['video_count'],
            "detected_subject_id":   subject.pk if subject else None,
            "detected_subject_name": subject.display_name if subject else None,
            "detected_exam_type":    exam_type,
            "already_imported":      bool(existing),
            "existing_pk":           existing.pk if existing else None,
            "subjects":              subjects,
        })


@method_decorator(coach_required, name='dispatch')
class YoutubePlaylistImportView(View):
    def post(self, request):
        from exams_app.models import Subject
        from tasks_app.services.youtube_playlist_importer import (
            extract_playlist_id, import_playlist,
            PlaylistNotFound, YouTubeAPIError,
        )
        body = _body(request)
        url = (body.get("url") or "").strip()
        pid = extract_playlist_id(url)
        if not pid:
            return JsonResponse({"error": "Geçerli bir YouTube oynatma listesi URL'si girin."}, status=400)

        subject_id = body.get("subject_id")
        exam_type  = body.get("exam_type", "").upper()

        if exam_type not in ("TYT", "AYT"):
            return JsonResponse({"error": "Sınav tipi TYT veya AYT olmalı."}, status=400)

        subject = None
        if subject_id:
            try:
                subject = Subject.objects.get(pk=int(subject_id))
            except (Subject.DoesNotExist, ValueError):
                return JsonResponse({"error": "Seçilen ders bulunamadı."}, status=400)

        was_update = False
        from tasks_app.models import YouTubePlaylist
        was_update = YouTubePlaylist.objects.filter(playlist_id=pid).exists()

        try:
            playlist = import_playlist(pid, subject, exam_type, request.user)
        except PlaylistNotFound:
            return JsonResponse({"error": "Bu playlist bulunamadı veya gizli."}, status=404)
        except YouTubeAPIError as exc:
            return JsonResponse({"error": str(exc)}, status=502)

        videos = [
            {"id": v.pk, "title": v.title, "duration": v.duration_min}
            for v in playlist.videos.order_by('position')
        ]
        return JsonResponse({
            "playlist_pk":      playlist.pk,
            "title":            playlist.title,
            "subject_id":       playlist.subject_id,
            "subject_display":  playlist.subject.display_name if playlist.subject else exam_type,
            "exam_type":        playlist.exam_type,
            "was_update":       was_update,
            "video_count":      len(videos),
            "videos":           videos,
        }, status=201)


@method_decorator(coach_required, name='dispatch')
class YoutubePlaylistListView(View):
    def get(self, request):
        from tasks_app.models import YouTubePlaylist
        qs = YouTubePlaylist.objects.select_related('subject').order_by('exam_type', 'title')
        playlists = [
            {
                "pk":              p.pk,
                "title":           p.title,
                "channel_title":   p.channel_title,
                "exam_type":       p.exam_type,
                "subject_display": p.subject.display_name if p.subject else "",
                "video_count":     p.videos.count(),
            }
            for p in qs
        ]
        return JsonResponse({"playlists": playlists})


@method_decorator(coach_required, name='dispatch')
class YoutubePlaylistVideosView(View):
    def get(self, request, pk: int):
        from tasks_app.models import YouTubePlaylist
        try:
            playlist = YouTubePlaylist.objects.get(pk=pk)
        except YouTubePlaylist.DoesNotExist:
            return JsonResponse({"error": "Playlist bulunamadı."}, status=404)
        videos = [
            {"id": v.pk, "title": v.title, "duration": v.duration_min}
            for v in playlist.videos.order_by('position')
        ]
        return JsonResponse({
            "playlist_pk":      playlist.pk,
            "title":            playlist.title,
            "subject_display":  playlist.subject.display_name if playlist.subject else "",
            "exam_type":        playlist.exam_type,
            "videos":           videos,
        })


@method_decorator(coach_required, name='dispatch')
class YoutubePlaylistDeleteView(View):
    def delete(self, request, pk: int):
        from tasks_app.models import YouTubePlaylist
        try:
            playlist = YouTubePlaylist.objects.get(pk=pk)
        except YouTubePlaylist.DoesNotExist:
            return JsonResponse({"error": "Playlist bulunamadı."}, status=404)
        playlist.delete()
        return JsonResponse({"deleted": pk})


# ── Excel export ──────────────────────────────────────────────────────────────

@method_decorator(coach_required, name='dispatch')
class ExportXlsxView(View):
    def get(self, request):
        try:
            student_id = int(request.GET["student_id"])
            hafta = date.fromisoformat(request.GET["hafta"])
        except (KeyError, ValueError):
            return HttpResponse("student_id ve hafta=YYYY-MM-DD gerekli", status=400)

        basi, sonu = week_svc.week_bounds(hafta)
        gorevler = tasks.week_for_student(request.user, student_id, basi, sonu)

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        RENK = {
            "konu_anlatimi": "FFDBEDFF",
            "soru_cozumu":   "FFFFF3CC",
            "tekrar":        "FFD4FFDA",
        }
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Hafta {basi}"

        # Header row: day names + date
        from datetime import timedelta
        for col, (label, delta) in enumerate(zip(GUNLER, range(7)), start=1):
            day = basi + timedelta(days=delta)
            cell = ws.cell(row=1, column=col, value=f"{label}\n{day.strftime('%d.%m')}")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="FFE8EBF0")
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 22
        ws.row_dimensions[1].height = 30

        # Group tasks by day-of-week
        by_dow = {i: [] for i in range(7)}
        for g in gorevler:
            dow = date.fromisoformat(g["tarih"]).weekday()
            by_dow[dow].append(g)

        max_tasks = max((len(v) for v in by_dow.values()), default=0)
        for row_offset in range(max_tasks):
            row = 2 + row_offset
            for col in range(7):
                task_list = by_dow[col]
                if row_offset < len(task_list):
                    g = task_list[row_offset]
                    detay_lines = " / ".join(d["aciklama"] for d in g["detaylar"] if d["aciklama"])
                    dur = f"⏱{g['ozel_sure_dk']}dk  " if g["ozel_sure_dk"] else ""
                    text = f"{g['ders_title'] or ''}\n{dur}{detay_lines}"
                    cell = ws.cell(row=row, column=col + 1, value=text.strip())
                    color = RENK.get(g["aktivite_tipi"], "FFFFFFFF")
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.font = Font(size=8)
                    cell.border = border
                else:
                    ws.cell(row=row, column=col + 1).border = border
            ws.row_dimensions[row].height = 40 if max_tasks else 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"hafta_{basi}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# ── Student views (read-only + completion) ────────────────────────────────────

@method_decorator(student_required, name='dispatch')
class StudentHaftaView(View):
    def get(self, request):
        return render(request, "student/tasks/hafta.html", {
            "student_id": request.user.id,
        })


@method_decorator(student_required, name='dispatch')
class StudentGorevListView(View):
    def get(self, request):
        try:
            hafta = date.fromisoformat(request.GET["hafta"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "hafta=YYYY-MM-DD gerekli"}, status=400)
        basi, sonu = week_svc.week_bounds(hafta)
        gorevler = tasks.week_for_own_student(request.user, basi, sonu)
        return JsonResponse({
            "hafta_basi": basi.isoformat(),
            "hafta_sonu": sonu.isoformat(),
            "gorevler": gorevler,
            "gunluk_toplamlar": week_svc.daily_totals(gorevler),
        })


@method_decorator(student_required, name='dispatch')
class StudentGorevCompleteView(View):
    def post(self, request, pk: int):
        body = _body(request)
        quality = body.get("quality")  # "easy" | "medium" | "hard" | None
        data = tasks.toggle_complete(request.user, pk, quality)
        if not data:
            return JsonResponse({"errors": "Bulunamadı"}, status=404)
        return JsonResponse({"ok": True, "gorev": data})


@method_decorator(student_required, name='dispatch')
class StudentGorevEditView(View):
    def put(self, request, pk: int):
        body = _body(request)
        data = tasks.student_update(
            request.user, pk,
            ozel_sure_dk=body.get("ozel_sure_dk"),
            aciklama=body.get("aciklama"),
        )
        if not data:
            return JsonResponse({"errors": "Bulunamadı veya düzenleme izni yok"}, status=404)
        return JsonResponse({"ok": True, "gorev": data})


@method_decorator(student_required, name='dispatch')
class StudentGorevReorderView(View):
    def post(self, request, pk: int):
        try:
            hedef_id = int(_body(request)["hedef_id"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "hedef_id gerekli"}, status=400)
        ok = tasks.student_swap_order(request.user, pk, hedef_id)
        return JsonResponse({"ok": ok}, status=200 if ok else 400)


@method_decorator(student_required, name='dispatch')
class StudentGorevCopyView(View):
    def post(self, request, pk: int):
        try:
            hedef = date.fromisoformat(_body(request)["hedef_tarih"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "hedef_tarih gerekli"}, status=400)
        yeni = tasks.student_copy_to_date(request.user, pk, hedef)
        if not yeni:
            return JsonResponse({"ok": False}, status=404)
        return JsonResponse({"id": yeni.id}, status=201)


# ── Student create + delete + reset ──────────────────────────────────────────

@method_decorator(student_required, name='dispatch')
class StudentGorevCreateView(View):
    def post(self, request):
        body = _body(request)
        try:
            tarih = date.fromisoformat(body["tarih"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "tarih gerekli"}, status=400)
        aktivite_tipi = body.get("aktivite_tipi", "tekrar")
        ders_title    = (body.get("ders_title") or "").strip() or None
        ozel_sure_dk  = body.get("ozel_sure_dk")
        aciklama      = (body.get("aciklama") or "").strip()
        detaylar = [{"aciklama": aciklama}] if aciklama else []
        grup = tasks.student_create(
            request.user,
            tarih=tarih, ders_title=ders_title, ozel_sure_dk=ozel_sure_dk,
            aktivite_tipi=aktivite_tipi, detaylar=detaylar,
        )
        return JsonResponse({"id": grup.id}, status=201)


@method_decorator(student_required, name='dispatch')
class StudentGorevDeleteView(View):
    def delete(self, request, pk: int):
        ok = tasks.student_hide_or_delete(request.user, pk)
        return JsonResponse({"ok": ok}, status=200 if ok else 404)


@method_decorator(student_required, name='dispatch')
class StudentResetView(View):
    def post(self, request):
        try:
            hafta = date.fromisoformat(_body(request)["week_start"])
        except (KeyError, ValueError):
            return JsonResponse({"errors": "week_start gerekli"}, status=400)
        basi, sonu = week_svc.week_bounds(hafta)
        deleted = tasks.reset_student_week(request.user, basi, sonu)
        return JsonResponse({"ok": True, "deleted": deleted})


@method_decorator(coach_required, name='dispatch')
class CoachStudentResetView(View):
    def post(self, request):
        body = _body(request)
        try:
            student_id = int(body["student_id"])
            hafta = date.fromisoformat(body["week_start"])
        except (KeyError, ValueError, TypeError):
            return JsonResponse({"errors": "student_id ve week_start gerekli"}, status=400)
        basi, sonu = week_svc.week_bounds(hafta)
        deleted = tasks.reset_student_week_by_coach(request.user, student_id, basi, sonu)
        return JsonResponse({"ok": True, "deleted": deleted})


# ── Student Excel export ──────────────────────────────────────────────────────

@method_decorator(student_required, name='dispatch')
class StudentExportXlsxView(View):
    def get(self, request):
        try:
            hafta = date.fromisoformat(request.GET["hafta"])
        except (KeyError, ValueError):
            return HttpResponse("hafta=YYYY-MM-DD gerekli", status=400)
        basi, sonu = week_svc.week_bounds(hafta)
        gorevler = tasks.week_for_own_student(request.user, basi, sonu)

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from datetime import timedelta

        GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        RENK = {
            "konu_anlatimi": "FFDBEDFF",
            "soru_cozumu":   "FFFFF3CC",
            "tekrar":        "FFD4FFDA",
        }
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook(); ws = wb.active; ws.title = f"Hafta {basi}"
        for col, (label, delta) in enumerate(zip(GUNLER, range(7)), start=1):
            day = basi + timedelta(days=delta)
            cell = ws.cell(row=1, column=col, value=f"{label}\n{day.strftime('%d.%m')}")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="FFE8EBF0")
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 22
        ws.row_dimensions[1].height = 30

        by_dow = {i: [] for i in range(7)}
        for g in gorevler:
            dow = date.fromisoformat(g["tarih"]).weekday()
            by_dow[dow].append(g)

        max_tasks = max((len(v) for v in by_dow.values()), default=0)
        for row_offset in range(max_tasks):
            row = 2 + row_offset
            for col in range(7):
                tl = by_dow[col]
                if row_offset < len(tl):
                    g = tl[row_offset]
                    details = " / ".join(d["aciklama"] for d in g["detaylar"] if d["aciklama"])
                    dur = f"⏱{g['ozel_sure_dk']}dk  " if g["ozel_sure_dk"] else ""
                    done = " ✓" if g.get("is_completed") else ""
                    cell = ws.cell(row=row, column=col + 1, value=f"{g['ders_title'] or ''}{done}\n{dur}{details}".strip())
                    cell.fill = PatternFill("solid", fgColor=RENK.get(g["aktivite_tipi"], "FFFFFFFF"))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.font = Font(size=8)
                    cell.border = border
                else:
                    ws.cell(row=row, column=col + 1).border = border
            ws.row_dimensions[row].height = 40

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="hafta_{basi}_benim.xlsx"'
        return response
