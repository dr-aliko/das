from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.template.loader import render_to_string

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from kocluk.constants import RENK_MAP
from kocluk.models import AktiviteTipi, Ogrenci
from kocluk.services import tasks, week as week_svc

User = get_user_model()

GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]


def weekly_html(koc: User, ogrenci_id: int, hafta_basi: date) -> str:
    _, hafta_sonu = week_svc.week_bounds(hafta_basi)
    gorevler = tasks.week_for_student(koc, ogrenci_id, hafta_basi, hafta_sonu)
    ogrenci = Ogrenci.objects.filter(koc=koc, pk=ogrenci_id).first()
    gunluk = week_svc.daily_totals(gorevler)
    return render_to_string(
        "kocluk/exports/hafta_html.html",
        {
            "ogrenci": ogrenci,
            "hafta_basi": hafta_basi,
            "hafta_sonu": hafta_sonu,
            "gunler": _build_days(gorevler, hafta_basi),
            "gunluk_toplamlar": {
                i: week_svc.format_minutes(v) for i, v in gunluk.items()
            },
            "RENK_MAP": RENK_MAP,
            "AktiviteTipi": AktiviteTipi,
        },
    )


def weekly_xlsx(koc: User, ogrenci_id: int, hafta_basi: date) -> bytes:
    _, hafta_sonu = week_svc.week_bounds(hafta_basi)
    gorevler = tasks.week_for_student(koc, ogrenci_id, hafta_basi, hafta_sonu)
    ogrenci = Ogrenci.objects.filter(koc=koc, pk=ogrenci_id).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Haftalık Program"

    header_font = Font(bold=True, size=11)
    for col_idx, (gun_adi, gun_date) in enumerate(
        _iter_days(hafta_basi), start=1
    ):
        day_gorevler = [
            g for g in gorevler if g["tarih"] == gun_date.isoformat()
        ]
        header_cell = ws.cell(row=1, column=col_idx, value=f"{gun_adi} {gun_date.strftime('%d.%m')}")
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[header_cell.column_letter].width = 45

        for row_idx, grup in enumerate(day_gorevler, start=2):
            label = _grup_label(grup)
            cell = ws.cell(row=row_idx, column=col_idx, value=label)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            color = _hex_fill(grup.get("aktivite_tipi"))
            if color:
                cell.fill = PatternFill("solid", fgColor=color.lstrip("#"))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── helpers ──────────────────────────────────────────────────────────────────

def _build_days(gorevler: list[dict], hafta_basi: date) -> list[dict]:
    from datetime import timedelta
    days = []
    for i in range(7):
        gun_date = hafta_basi + timedelta(days=i)
        day_tasks = [g for g in gorevler if g["tarih"] == gun_date.isoformat()]
        days.append({"gun": GUNLER[i], "tarih": gun_date, "gorevler": day_tasks})
    return days


def _iter_days(hafta_basi: date):
    from datetime import timedelta
    for i in range(7):
        yield GUNLER[i], hafta_basi + timedelta(days=i)


def _grup_label(grup: dict) -> str:
    aktivite = dict(AktiviteTipi.choices).get(grup.get("aktivite_tipi", ""), "")
    parts = [f"{aktivite} - {grup['ders_title']} ({grup['ozel_sure_dk']} dk)"]
    for d in grup.get("detaylar", []):
        sure = f" ({d['sure_bilgisi']})" if d.get("sure_bilgisi") else ""
        parts.append(f"• {d['aciklama']}{sure}")
    return "\n".join(parts)


def _hex_fill(aktivite_tipi: str | None) -> str:
    return RENK_MAP.get(aktivite_tipi, "")
